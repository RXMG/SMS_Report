#!/usr/bin/env python
# coding: utf-8

# In[1]:


import pygsheets
import pandas as pd
import numpy as np
import datetime as dt
import time
import filepath
import pandas as pd
import glob
import os
import numpy as np
import send_email 
import pygsheets
import infrastructure
import datetime
import math
import datetime
from datetime import timedelta  
import re
import warnings
import statistics
from datetime import date
from calendar import monthrange
from dateutil.relativedelta import relativedelta
import xlsxwriter
import send_email
from colorama import Fore, Style
import matplotlib.pyplot as plt
import seaborn as sns
import datetime as dt
import pytz
from sklearn.linear_model import LinearRegression
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import FormulaRule

timezone = pytz.timezone('US/Pacific')
datetime.datetime.now(timezone)


# In[2]:


## zero revenue offer 


# In[3]:


sms =pd.read_csv(filepath.output_folder+'SS_LC_merged_data.csv')
#sms = pd.read_csv('SS_LC_merged_data.csv')
data = infrastructure.transform_sms_df(sms)
data['Hitpath Offer ID'] = data['Hitpath Offer ID'].astype(str).str.split('.',expand = True)[0]
data = data[data['Hitpath Offer ID'].isna() == False]
all_send_strategy_df = data.copy()


# In[4]:


all_send_strategy_df


# In[5]:


offer_sms = infrastructure.get_smartsheet('offers_sms')
offer_sms = offer_sms[offer_sms['Hitpath Offer ID'].isna() == False]
#add rx rep
offer_sms.rename(columns = {'RX Rep':'AM'}, inplace = True)
offer_sms['Hitpath Offer ID'] = offer_sms['Hitpath Offer ID'].astype(str).str.split('.',expand = True)[0]
current_date = dt.date.today()  # Get the current date
max_date = current_date - dt.timedelta(days=1)
last_7_days = max_date - dt.timedelta(days=7)
zero_ot_email_body = ""
last_7d_df = all_send_strategy_df[(all_send_strategy_df['Date'].dt.date< max_date ) & (all_send_strategy_df['Date'].dt.date>=  last_7_days) ]
tableau_link = "https://tableau.rxmg.app/views/RXMG-SMSReports/0Drops-PT?:showAppBanner=false&:display_count=n&:showVizHome=n&:origin=viz_share_link"
zero_ot_email_body += "<h3>* Here are Pipeline Testing Offers with $0 Revenue, please check whether there's tracking issue:</h3>"
zero_ot_email_body += f"<p>Tableau: <a href=\"{tableau_link}\">LINK</a></p>"
last_7d_df['Hitpath Offer ID'] = last_7d_df['Hitpath Offer ID'].astype(str).str.split('.',expand = True)[0]
temp = last_7d_df.loc[last_7d_df['Send Strategy'] == 'PT',].groupby('Hitpath Offer ID')['Revenue'].sum()
zero_drops = temp[temp == 0].reset_index(name = 'Revenue')
if len(zero_drops) != 0 : 
    zero_drops = zero_drops.merge(offer_sms[['Hitpath Offer ID','Scheduling Name','AM']], how = 'left',left_on = 'Hitpath Offer ID',right_on = 'Hitpath Offer ID')
    for i in zero_drops['Hitpath Offer ID'].unique(): 
        zero_ot_email_body += f'<p><b>Offer: {scheduling_name}</b></p>'
        zero_ot_email_body += '<table border="1" style="border-collapse: collapse;">'
        zero_ot_email_body += '<tr><th>Date</th><th>DP&Pub</th></tr>'
        scheduling_name = zero_drops[zero_drops['Hitpath Offer ID']==i]['Scheduling Name'].values[0]
        
        izero_drops = last_7d_df.loc[last_7d_df['Hitpath Offer ID'] == i, ] 
        for j in range(0,len(izero_drops)): 
            temp_df = izero_drops.iloc[j,:]
            date_temp_df =temp_df['Date'].strftime("%Y-%m-%d")
            dp_pub_temp_df = temp_df['DP&Pub']
            zero_ot_email_body += f"<tr><td>{date_temp_df}</td><td>{dp_pub_temp_df}</td></tr>"
    zero_ot_email_body += '</table>'
else: 
    zero_ot_email_body += "<p><b>There's no $0 Revenue pipeline testing offers in the last 7 days</b></p>"


# In[6]:


zero_ot_email_body


# In[7]:


## New part
df = all_send_strategy_df[~all_send_strategy_df['Send Strategy'].isin(['PT','RT','IT','OT','W','MI','Mi','AR'])]
#df = all_send_strategy_df[all_send_strategy_df['Send Strategy']=='P']
df_sorted = df.copy().sort_values('Date').reset_index(drop=True)


# In[8]:


scheduleMaster = df_sorted.groupby(['Hitpath Offer ID','Date']).agg({'Revenue':'sum','Clicks':'sum','Delivered':'sum','Jump Page Clicks':'sum','Opportunity Cost':'sum','Optout':'sum','Cost':'sum'}).reset_index()
scheduleMaster = scheduleMaster.merge(offer_sms[['Hitpath Offer ID','Scheduling Name','AM', 'Status']], how = 'left',left_on = 'Hitpath Offer ID',right_on = 'Hitpath Offer ID')
scheduleMaster['Text CTR']= scheduleMaster['Clicks']/scheduleMaster['Delivered']
scheduleMaster['eCPM']=scheduleMaster['Revenue']*1000/scheduleMaster['Delivered']
scheduleMaster['JP CTR']=scheduleMaster['Jump Page Clicks']/scheduleMaster['Delivered']
scheduleMaster['optout rate']=scheduleMaster['Optout']/scheduleMaster['Delivered']
scheduleMaster['Text EPC'] = scheduleMaster['Revenue']/scheduleMaster['Clicks']
scheduleMaster['JP EPC']= scheduleMaster['Revenue']/scheduleMaster['Jump Page Clicks']
scheduleMaster = scheduleMaster[scheduleMaster['Status'] == 'Live']
scheduleMaster.replace([np.inf, -np.inf], np.nan, inplace=True)
scheduleMaster.fillna(0, inplace=True)


# In[9]:


scheduleMaster


# In[10]:


offer_ids = scheduleMaster['Hitpath Offer ID'].unique()
offer_ids


# In[11]:


test = scheduleMaster[scheduleMaster['Hitpath Offer ID'] == '12610']
test.tail(10)


# In[12]:


def linear_reg(n_drops, 
               start_date, 
               end_date, 
               offer_id, 
               metrics, 
               eCPM_drop_pct, 
               scheduleMaster, 
               result_list): 
    specific_offer = scheduleMaster[scheduleMaster['Hitpath Offer ID'] == offer_id]
    sorted_offer = specific_offer.sort_values('Date').reset_index(drop=True)
    filtered_offer = sorted_offer[sorted_offer['Date'] >= start_date]
    last_n_data = filtered_offer.tail(n_drops)

    avg_eCPM = last_n_data['Revenue'].sum() * 1000 / last_n_data['Delivered'].sum()
    avg_text_CTR= last_n_data['Clicks'].sum()/last_n_data['Delivered'].sum()
    avg_jump_CTR = last_n_data['Jump Page Clicks'].sum()/last_n_data['Delivered'].sum()

    if last_n_data['Date'].max() >= end_date and len(last_n_data['Date']) == n_drops:
        y_ecpm = last_n_data['eCPM'].values

        X = np.arange(len(y_ecpm)).reshape(-1, 1)
        model_ecpm = LinearRegression()
        model_ecpm.fit(X, y_ecpm)
        
        slope_ecpm = model_ecpm.coef_[0]
        intercept_ecpm = model_ecpm.intercept_
        y_pred = model_ecpm.predict(X)
        
        min_value = 0  # Minimum desired value
        offset = max(0, min_value - np.min(y_pred))
        
        y_pred_adjusted = y_pred + offset

        percent_diffs_pred = np.diff(y_pred_adjusted) / y_pred_adjusted[:-1]
        avg_eCPM_drop_pct = np.mean(percent_diffs_pred)
        
        result_dict = {}
        alert = False
        note = ""
        if slope_ecpm < 0 and avg_eCPM_drop_pct < eCPM_drop_pct:
            alert = True
            print(f"Offer_ID: {offer_id} \n{y_ecpm} \n{percent_diffs_pred}")
            for metric, reason in metrics.items():
                y_current_metric = last_n_data[metric].values
                model_metric = LinearRegression()
                model_metric.fit(X, y_current_metric)
                slope_metric = round(model_metric.coef_[0],2)
                cor = np.corrcoef(y_ecpm, y_current_metric)[0, 1]
                cor_str = corr_strength(cor)
                if slope_metric < 0 and cor > 0.5:
                    if len(note) == 0:
                        note += f"Potential problem:"
                    note += f" {reason}"
        if alert: 
            result_dict['Hitpath Offer ID'] = offer_id
            result_dict['Scheduling Name'] = specific_offer.iloc[0]['Scheduling Name']
            result_dict['AM'] = specific_offer.iloc[0]['AM']
            result_dict['avg_eCPM'] = to_dollar(avg_eCPM)
            result_dict['avg_eCPM_drop_pct'] = to_pct(avg_eCPM_drop_pct) 
            result_dict['avg_Text_CTR'] = to_pct(avg_text_CTR)
            result_dict['avg_Jump_CTR'] = to_pct(avg_jump_CTR)
            result_dict['note'] = note
            result_list.append(result_dict)

def to_pct(n):
    if isinstance(n, (int, float)):
        return "{:,.2f}%".format(n * 100)
    else:
        return n

def to_dollar(n):
    if isinstance(n, (int, float)):
        return "${:,.2f}".format(n)
    else: 
        return n


def corr_strength(correlation):
    if correlation == 1:
        explanation = "Perfect positive"
    elif correlation > 0.7:
        explanation = "Strong positive"
    elif correlation > 0.5:
        explanation = "Moderate positive"
    elif correlation > 0:
        explanation = "Weak positive"
    elif correlation == 0:
        explanation = "No linear relationship"
    elif correlation > -0.5:
        explanation = "Weak negative"
    elif correlation > -0.7:
        explanation = "Moderate negative"
    elif correlation > -1:
        explanation = "Strong negative"
    else:  # correlation == -1
        explanation = "Perfect negative"
    return explanation


# In[13]:


def avg_less_than(
               start_date,
               avg_low_threshold,
               offer_id, 
               scheduleMaster,      
               result_list): 
    specific_offer = scheduleMaster[scheduleMaster['Hitpath Offer ID'] == offer_id]
    sorted_offer = specific_offer.sort_values('Date').reset_index(drop=True)
    last_n_data = sorted_offer[sorted_offer['Date'] >= start_date]
    last_n_avg_eCPM = last_n_data['Revenue'].sum() * 1000 / last_n_data['Delivered'].sum()
    last_n_avg_text_CTR= last_n_data['Clicks'].sum() / last_n_data['Delivered'].sum()
    last_n_avg_jp_CTR = last_n_data['Jump Page Clicks'].sum() /last_n_data['Delivered'].sum()   
    
    result_dict = {}
    if last_n_avg_eCPM <= 7:
        result_dict['Hitpath Offer ID'] = offer_id
        result_dict['Scheduling Name'] = specific_offer.iloc[0]['Scheduling Name']
        result_dict['AM'] = specific_offer.iloc[0]['AM']
        result_dict['avg_eCPM'] = to_dollar(last_n_avg_eCPM)
        result_dict['avg_Text_CTR'] = to_pct(last_n_avg_text_CTR)
        result_dict['avg_Jump_CTR'] = to_pct(last_n_avg_jp_CTR)
        result_dict['number_of_drops'] = len(last_n_data)
        result_dict['note'] = ''
        result_list.append(result_dict)


# In[14]:


metrics = {'Text CTR': 'Content', 
           'JP CTR': 'Jump Page'}


n_drops = 7 
down_trending_start_date = datetime.datetime.now() - timedelta(days=28)
down_trending_end_date =  datetime.datetime.now() - timedelta(days=14)

eCPM_drop_pct = -0.2
down_trending_result = []
for offer_id in offer_ids:
    linear_reg(n_drops, down_trending_start_date, down_trending_end_date, offer_id, metrics, eCPM_drop_pct, scheduleMaster, down_trending_result)

low_average_result = []
low_avg_start_date = datetime.datetime.now() - timedelta(days=14)
avg_low_threshold = 7
for offer_id in offer_ids:
    avg_less_than(low_avg_start_date, avg_low_threshold, offer_id, scheduleMaster, low_average_result)

down_trending_column_name = ['Hitpath Offer ID', 'Scheduling Name', 'AM', 'avg_eCPM', 'avg_Text_CTR', 'avg_Jump_CTR', 'avg_eCPM_drop_pct', 'note']   
low_average_column_name = ['Hitpath Offer ID', 'Scheduling Name', 'AM', 'avg_eCPM', 'avg_Text_CTR', 'avg_Jump_CTR', 'note', 'number_of_drops']   
down_trending_df = pd.DataFrame(down_trending_result, columns = down_trending_column_name)
low_average_df = pd.DataFrame(low_average_result, columns = low_average_column_name)


# In[15]:


#upComing_drops = pd.read_excel('./SMS_upcoming_schedule.xlsx', sheet_name='SMS Upcoming Schedule')
upComing_drops = pd.read_excel(filepath.output_folder + 'SMS_upcoming_schedule.xlsx', sheet_name='SMS Upcoming Schedule')
upComing_drops['Hitpath Offer ID'] = upComing_drops['Hitpath Offer ID'].astype(str).str.split('.',expand = True)[0]
upComing_drops = upComing_drops[~upComing_drops['Send Strategy'].isin(['PT','RT','IT','OT','W','MI','Mi','AR'])]
upComing_drops = upComing_drops[upComing_drops['Hitpath Offer ID'].isin(low_average_df['Hitpath Offer ID'])]
upComing_drops = upComing_drops[upComing_drops['Date'] > datetime.datetime.now()] 
unComing_offer_counts = upComing_drops['Hitpath Offer ID'].value_counts().reset_index()
unComing_offer_counts.columns = ['Hitpath Offer ID', 'uncoming_drops_count']

low_average_df = low_average_df.merge(unComing_offer_counts, on='Hitpath Offer ID', how='left')
low_average_df['uncoming_drops_count'] = low_average_df['uncoming_drops_count'].fillna(0).astype(int)


# In[16]:


low_average_html_content = ""

low_average_html_content = f"<h3>*Offer that had avg eCPM less than $7 within last 14 days(P&CT drops):</h3>"
if len(low_average_df):
    low_average_html_content += '<table border="1" style="border-collapse: collapse;">'
    low_average_html_content += '<tr><th>Scheduling Name</th><th>AM</th><th>Avg eCPM</th><th>Number of past drops</th><th>Number of upcoming drops</th></tr>'
    for index, row in low_average_df.iterrows():
        scheduling_name = row['Scheduling Name']
        am = row['AM']
        avg_ecpm = row['avg_eCPM']
        number_of_drops = row['number_of_drops']
        number_of_upcoming_drops = row['uncoming_drops_count']
        low_average_html_content += f'<tr><td>{scheduling_name}</td><td>{am}</td><td>{avg_ecpm}</td><td>{number_of_drops}</td><td>{number_of_upcoming_drops}</td></tr>'
    low_average_html_content += '</table>'
else:
    low_average_html_content += "<p>No offers found.</p>"


# In[17]:


down_trending_html_content = ""

down_trending_html_content = f"<h3>*Offers downtrending in latest {n_drops} distinct days with projected average dropped more than {to_pct(eCPM_drop_pct)}(P&CT drops):</h3>"
if len(down_trending_df):
    down_trending_html_content += '<table border="1" style="border-collapse: collapse;">'
    down_trending_html_content += '<tr><th>Scheduling Name</th><th>AM</th><th>Avg eCPM</th><th>Avg Drop Pct</th><th>Note</th></tr>'
    for index, row in down_trending_df.iterrows():
        scheduling_name = row['Scheduling Name']
        am = row['AM']
        avg_ecpm = row['avg_eCPM']
        avg_eCPM_drop_pct = row['avg_eCPM_drop_pct']
        note = row['note']
        down_trending_html_content += f'<tr><td>{scheduling_name}</td><td>{am}</td><td>{avg_ecpm}</td><td>{avg_eCPM_drop_pct}</td><td>{note}</td></tr>'
    down_trending_html_content += '</table>'
else:
    down_trending_html_content += "<p>No down trending offer found.</p>"


# In[18]:


file_path = 'Offer Alert - {}.xlsx'.format(date.today().strftime("%m_%d_%Y"))

merged_df = pd.merge(low_average_df, down_trending_df, on=['Hitpath Offer ID', 'Scheduling Name', 'AM'], suffixes=('_df1', '_df2'), how='outer')
merged_df['low_avg'] = merged_df.apply(lambda row: True if pd.notna(row['avg_eCPM_df1']) else False, axis=1)
merged_df['down_trending'] = merged_df.apply(lambda row: True if pd.notna(row['avg_eCPM_df2']) else False, axis=1)
merged_df['note'] = merged_df['note_df1'].combine_first(merged_df['note_df2'])
final_df = merged_df[['Hitpath Offer ID', 'Scheduling Name', 'AM', 'low_avg', 'down_trending', 'note']]

final_df.rename(columns={'low_avg': f'low_avg_{avg_low_threshold}'}, inplace=True)


# In[19]:


def write_date_to_excel(df_sorted, offer_sms, offer_ids, start_date, writer, name, n_drops= -1):
    offer_data = df_sorted[df_sorted['Hitpath Offer ID'].isin(offer_ids)]
    if n_drops != -1:
        last_seven_date = offer_data['Date'].drop_duplicates().nlargest(n_drops) 
        last_seven_offer_data = offer_data[offer_data['Date'].isin(last_seven_date)]
    else:
        last_seven_offer_data = offer_data[offer_data['Date'] >= start_date]
        
    last_seven_offer_data.fillna(0, inplace=True)
    last_seven_offer_data = last_seven_offer_data[['Hitpath Offer ID', 'Date', 'SC_DP&Pub', 'eCPM', 'Clicks', 'Delivered', 'Revenue', 'Jump Page Clicks']]
    result = last_seven_offer_data.merge(offer_sms[['Hitpath Offer ID','Scheduling Name','AM']], how = 'left',left_on = 'Hitpath Offer ID',right_on = 'Hitpath Offer ID')
    result['Text CTR']= result['Clicks']/result['Delivered']
    result['Text CTR'] = result['Text CTR'].apply(to_pct)
    result['JP CTR']=result['Jump Page Clicks']/result['Delivered']
    result['Text EPC'] = result['Revenue']/result['Clicks']
    result['Text EPC']= result['Text EPC'].apply(to_dollar) 
    result['JP EPC']= result['Revenue']/result['Jump Page Clicks']
    result['JP EPC']= result['JP EPC'].apply(to_dollar) 
    result['JP CTR'] = result['JP CTR'].apply(to_pct)
    result['eCPM']= result['eCPM'].apply(to_dollar) 
    sorted_last_seven_offer_data = result.sort_values(by=['Hitpath Offer ID','Date'])
    sorted_last_seven_offer_data = sorted_last_seven_offer_data[['Date','Hitpath Offer ID','Scheduling Name','AM', 'SC_DP&Pub','eCPM','Delivered', 'Revenue','Clicks','Text CTR','Text EPC','Jump Page Clicks','JP CTR','JP EPC']]
    sorted_last_seven_offer_data.to_excel(writer, index=False, sheet_name=name)


# In[20]:


with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    final_df.to_excel(writer, index=False, sheet_name='summary')
    write_date_to_excel(df_sorted, offer_sms, low_average_df['Hitpath Offer ID'], low_avg_start_date, writer, f'low_avg_$7', -1)
    write_date_to_excel(df_sorted, offer_sms, down_trending_df['Hitpath Offer ID'], down_trending_start_date, writer, 'down_trending_offers', n_drops)


# In[21]:


workbook = openpyxl.load_workbook(file_path)
worksheet = workbook['summary']
red_font = Font(color="FF0000")
for row in range(2, worksheet.max_row + 1):  # Adjusting for header and 1-based index
    cell = worksheet.cell(row=row, column=4)  # Assuming the 'avg less than 7' column is the 5
    cell_1 = worksheet.cell(row=row, column=5)  # Assuming the 'avg less than 7' column is the 5
    if cell.value is True:
        cell.font = red_font
    if cell_1.value is True:
        cell_1.font = red_font

# Save the workbook
workbook.save(file_path)


# In[22]:


toaddr = ['offernotices@rxmg.com','lili@rxmg.com','nathan@rxmg.com','nina@rxmg.com','n.ohashi@rxmg.com']
#toaddr = ['nina@rxmg.com']
today = date.today().strftime("%m_%d_%Y")
subject_line = f"SMS Team Offer Alert Report - {today}"
html_content = ""
html_content += """\
<html>
  <body>
   <p>Hi Team: </p>
   <p>Please find the attached Offer Alert Report and optimize your scheduling.</p>
"""

html_content +="<br>"

html_content +=zero_ot_email_body
html_content +="<br>"
html_content +=low_average_html_content
html_content +="<br>"
html_content +=down_trending_html_content


html_content += """\
  </body>
</html>
"""

for i in toaddr:
    send_email.send_email([file_path],subject_line,html_content,i, 'html')


# In[ ]:




