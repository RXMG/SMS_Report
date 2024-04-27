import sys
# import builtins
# sys.path.append(builtins.viper_local_schedule_codes_path)
import os
import shutil
from os import listdir
from os.path import isfile, join
# filepath_file = builtins.viper_local_schedule_codes_path + '/filepath.py'
# infrastructure_file = builtins.viper_local_schedule_codes_path + '/infrastructure.py'
# dst_path = os.getcwd()
# shutil.copy(filepath_file, dst_path)
# shutil.copy(infrastructure_file, dst_path)
# import infrastructure
import infrastructure
del sys.modules["infrastructure"]
import infrastructure
from io import BytesIO
import pygsheets
import pandas as pd
import numpy as np
import datetime as dt
import time
import filepath
del sys.modules["filepath"]
import filepath
import infrastructure
import math
import datetime
from datetime import timedelta  
# import emoji
import re
import warnings
import statistics
from datetime import date
import itertools
import random
os.environ['PYGAME_HIDE_SUPPORT_PROMPT'] = "hide"
import pygame
import numpy as np
import pandas as pd
import signal
import json
#pip install google-cloud-storage
from google.cloud import storage
import filepath
import infrastructure
from scipy.stats import percentileofscore

# import matplotlib.pyplot as plt
# import seaborn as sns
from openpyxl import load_workbook

warnings.filterwarnings('ignore', '.*invalid value encountered in scalar divide*', )
warnings.filterwarnings('ignore', '.*Comparison of Timestamp*', )
warnings.filterwarnings('ignore', '.*apply openpyxl.*', )
warnings.filterwarnings('ignore', '.*Specify dtype option on import.*', )
warnings.filterwarnings('ignore', '.*empty string*', )
warnings.filterwarnings('ignore', '.*invalid value encountered in double_scalars*', )

pd.options.mode.chained_assignment = None  # default='warn'
pd.set_option('display.max_columns',50)
pd.set_option('display.max_rows',50)

def hiss():
    try:
        pygame.mixer.init()
        pygame.mixer.music.load("snake-hissing-6092.mp3")
        pygame.mixer.music.play()
    except:
        print('HISS!')
    

def rattle():
    try:
        pygame.mixer.init()
        pygame.mixer.music.load("rattle.mp3")
        pygame.mixer.music.play()
    except:
        print('RATTLE!')
        
def get_emit():
    print('\nLoading EMIT...')
    emit = infrastructure.get_publisher()
    emit['Affiliate ID'] = emit['PUBID'].astype(str).str.strip()
    
    emit['Affiliate ID'] = emit['Affiliate ID'].apply(lambda x: "NaN" if x == '' else x)
    print('EMIT Loaded!')
    return emit

def get_offers():
    print('\nLoading Offers Sheet...')
    offers = infrastructure.get_smartsheet('offers_sms')
    offers.dropna(subset=['Hitpath Offer ID'],inplace=True)
    try:
        offers['Hitpath Offer ID'] = offers['Hitpath Offer ID'].astype('str').str.split('.',expand = True)[0]
    except ValueError as e:
        hiss()
        print('Error Loading the Offers Sheet: {}'.format(e))
        
    # add offer wall into offers sheet
    offerwall = infrastructure.get_smartsheet('ow_sms')
    offerwall = offerwall[['Hitpath Offer ID','Vertical','Status','Offer List','Offer Wall Link']]
    offerwall = offerwall.rename(columns={ 'Offer List': 'Scheduling Name','Offer Wall Link':'Redirect Link'})
    offers = pd.concat([offers,offerwall],ignore_index=True)

    print('Offers Sheet Loaded!')
    offers = offers.set_index('Hitpath Offer ID')
    offers = offers[~offers.index.duplicated(keep='first')]
    set_offers(offers)
    return offers

def set_offers(offers2):
    global offers
    offers=offers2
    
def get_offers_loaded():
    global offers
    return offers

def calculate_oppo_ctr(df):
    df['opportunity CTR'] = df['Opportunity Clicks'] / df['Delivered']
    #df['opportunity CTR Normalized']=(df['opportunity CTR']-df['opportunity CTR'].min())/(df['opportunity CTR'].max()-df['opportunity CTR'].min())
    opportunity_ctr_series = df[(df['opportunity CTR']!=np.inf)&(df['opportunity CTR'].isna() == False)]['opportunity CTR']
    df['opportunity CTR Normalized']= percentileofscore(opportunity_ctr_series, df['opportunity CTR'], kind='weak') / 100
    df['opportunity eCPM'] = df['Opportunity Cost'] * 1000 / df['Delivered']
    #df['opportunity eCPM Normalized']=(df['opportunity eCPM']-df['opportunity eCPM'].min())/(df['opportunity eCPM'].max()-df['opportunity eCPM'].min())
    opportunity_eCPM_series =  df[(df['opportunity eCPM']!=np.inf)&(df['opportunity eCPM'].isna() == False)]['opportunity eCPM']
    df['opportunity eCPM Normalized']=percentileofscore(opportunity_eCPM_series, df['opportunity eCPM'], kind='weak') / 100 
    df['opportunity CTR50'] = df['opportunity CTR Normalized']*0.6 + df['opportunity eCPM Normalized']*0.4
    return df

def get_lexi_fast():
    #This will load the most recently modified version of Lexi located in your downloads folder
    
    if os.environ.get('DOCKER_SCRIPT_RUNNING'):
        # In SMS Team, we don't store data in the Cloud 
        from google.cloud import storage

        print('Checking viper-bucket')
        # viper-bucket/Lexi 3.0 - 2023-12-04T142046.393.xlsx
        bucket_name = "viper-bucket"
        #auth_file_name = 'Lexi.xlsx'
        auth_file_name = 'SS_LC_merged_data.csv'
        # folder_path = 'Authentication_and_Data/'
        local_auth_directory = '/tmp/' 

        # os.makedirs(local_auth_directory, exist_ok=True)

        lexi_path = local_auth_directory + auth_file_name

        storage_client = storage.Client()

        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(auth_file_name)
#         print('blob is',blob)

        blob.download_to_filename(lexi_path)

    elif os.environ.get('DOCKER_RUNNING'):
        lexi_download_path='host_downloads/'
        
        #lexi_files = [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f)) 
        #              & ('Lexi 3.0' in f) & ~("Year" in f)]
        lexi_files =  [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f))& ('SS_LC_merged_data' in f)]
        lexi_paths = ["{}/{}".format(lexi_download_path,lexi_version) for lexi_version in lexi_files]
        times = [os.path.getmtime(path) for path in lexi_paths]
        # times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
        most_recent_lexi = lexi_files[times.index(max(times))]
        lexi_path = "{}/{}".format(lexi_download_path,most_recent_lexi)
    else:
        lexi_download_path=filepath.downloadpath
    
        #lexi_files = [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f)) 
        #              & ('Lexi 3.0' in f) & ~("Year" in f)]
        lexi_files = [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f))& ('SS_LC_merged_data' in f)]
        lexi_paths = ["{}/{}".format(lexi_download_path,lexi_version) for lexi_version in lexi_files]
        times = [os.path.getmtime(path) for path in lexi_paths]
        # times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
        most_recent_lexi = lexi_files[times.index(max(times))]
        lexi_path = "{}/{}".format(lexi_download_path,most_recent_lexi)

    print('\nLoading Lexi Data...')
#     start = time.time()
    """ 
    wb = load_workbook(filename=lexi_path, read_only=True)

    ws = wb['Sheet1']

    # Read the cell values into a list of lists
    data_rows = []

    for row in ws.iter_rows(max_row = 1):
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    for row in ws.iter_rows(min_row = ws.max_row - 160000): # this should grab approximately the last 400 days worth of data
        data_cols = []
        for cell in row:
            data_cols.append(cell.value)
        data_rows.append(data_cols)

    # Transform into dataframe
    lexi = pd.DataFrame(data_rows)
    
    
    new_header = lexi.iloc[0] #grab the first row for the header
    lexi = lexi[1:] #take the data less the header row
    lexi.columns = new_header #set the header row as the df header
    """
    lexi = pd.read_csv(lexi_path)
    lexi = infrastructure.transform_sms_df(lexi) # transform sms data into lexi format 
    float64_cols = ['Sent',
     'Opens',
     'Open Rate',
     'Clicks',
     'Click Rate',
     'Contacts Lost',
     'Adjusted Clicks',
     'Adjusted Click Rate',
     'Revenue',
     'Revenue CPM (eCPM)',
     'Conversions',
     'Cost CPM',
     'Cost per send',
     'Net Revenue',
     'overall_inbox_percent',
     'aol_inbox_percent',
     'google_inbox_percent',
     'yahoo_inbox_percent',
     'outlook_inbox_percent',
     'global_isps_inbox_percent',
     'Vertical Gap',
     'Offer Gap',
     'Content Gap',
     'CTR50',
     'CTR25',
     'CTR10',
     'Opportunity Cost',
     'Vertical CTR Cost Ratio',
    ]
    float64_cols = ['Sent',                                                                       
                    'Revenue'  ,                                
                    'Jump Page Clicks' ,                        
                    'Delivered' ,                              
                    'Not_Delivered',                            
                    'Optout',                                   
                    'Clicks',                                   
                    'Cost',                                     
                    'Revenue CPM (eCPM)', 
                    'CTR',                                      
                    'Gross Profit',                             
                    'Margin',                                   
                    'RPU',                                      
                    'eCPM',                                     
                    'eCPM Ratio',                               
                    'CTR Normalized',                           
                    'eCPM Normalized',                          
                    'CTR50',                                    
                    'Profit', 
                    'Dataset_Agg_30D_eCPM',                     
                    'Opportunity Cost',      
                    'Vertical Gap',
                    'Offer Gap'
    ]

    for col in float64_cols:
        lexi[col] = lexi[col].astype('float64')
    lexi['Delivered'] = lexi['Delivered'].astype('int64')
    #lexi['Hitpath Offer ID'] = lexi['Hitpath Offer ID'].astype('int64')
    lexi['Date'] = lexi['Date'].astype('datetime64[ns]')
    
    lexi['eCPM'] = lexi['Revenue']*1000 / lexi['Delivered']
    lexi['CTR'] = lexi['Clicks']*1000 / lexi['Delivered']
    lexi = lexi[~lexi.duplicated(subset=['SC_DP&Pub','Affiliate ID','Hitpath Offer ID','Date','Segments','Revenue','Delivered','Send Strategy'])]
    lexi['Hitpath Offer ID'] = lexi['Hitpath Offer ID'].astype('str').str.split('.',expand = True)[0]
    
    #Deal with Duplicate Affiliate IDs for different DP&Pubs
    #lexi.loc[ lexi['DP&Pub']=='SC.FHA_460398', 'Affiliate ID'] = '160398'
    #lexi.loc[ lexi['DP&Pub']=='SC.RF_460398', 'Affiliate ID'] = '260398'
    #lexi.loc[ lexi['DP&Pub']=='LPG.FHA_461128', 'Affiliate ID'] = '161128'
    #lexi.loc[ lexi['DP&Pub']=='LPG.RF_461128', 'Affiliate ID'] = '261128'
    
    #lexi['Conversions'].fillna(0,inplace=True)
    try:
        offers = get_offers_loaded() #for proper merging
    except:
        offers = get_offers()
    
    

    lexi = pd.merge(lexi,offers[['Day Restrictions','Payout Type','Scheduling Name','Status','Paused Pubs']],left_on='Hitpath Offer ID',right_on=offers.index, how='left')
    temp1= lexi.groupby(['Segments','Opportunity Cost Send Strategy','Date']).agg({'Revenue':'sum','Clicks':'sum','Delivered':'sum'}).reset_index()
    temp1[['rolling Revenue','rolling Clicks','rolling Delivered']] = temp1.groupby('Segments').shift(1).rolling(30, min_periods=5)[['Revenue','Clicks','Delivered']].sum().reset_index(drop=True)
    temp1['Dataset_Agg_30day_ctr'] = temp1['rolling Clicks'] / temp1['rolling Delivered']
    temp1['Dataset_Agg_30day_ecpm'] = temp1['rolling Revenue'] * 1000/ temp1['rolling Delivered']
    dataset_agg_ctr =  temp1[['Segments','Date','Opportunity Cost Send Strategy','Dataset_Agg_30day_ctr','Dataset_Agg_30day_ecpm']]
    lexi = lexi.merge(dataset_agg_ctr, how = 'left')
    lexi['Opportunity Clicks'] = lexi['Clicks'] - lexi['Dataset_Agg_30day_ctr'] * lexi['Delivered']
    lexi['Opportunity Cost'] = lexi['Revenue'] - lexi['Dataset_Agg_30day_ecpm'] * lexi['Delivered'] /1000  # revenue  - last 30 days average revenue in that drop 
    lexi['opportunity CTR'] = lexi['Opportunity Clicks'] / lexi['Delivered']
    lexi['opportunity eCPM'] = lexi['Opportunity Cost'] * 1000 / lexi['Delivered']
    lexi = calculate_oppo_ctr(lexi)

#     lexi['Affiliate ID'] = lexi['Affiliate ID'].astype(float).astype(int).astype(str) #sometimes Lexi loads with varying dtypes
    
    lexi_raw = lexi.copy()
    
    #segment_conditions = [
    #    lexi['Segment'].str.contains('120').fillna(False),
    #    lexi['Segment'].str.contains('60').fillna(False),
    #    lexi['Segment'].str.contains('7').fillna(False),
    #    lexi['Segment'].str.contains('O30').fillna(False),
    #    lexi['Segment'].str.contains('O1').fillna(False),
    #        ]

    #segment_choices = ['A120', 'A60', 'A7', 'O30', 'O1']
    #lexi['Super Segment'] = np.select(segment_conditions, segment_choices, default='Other')

    
    
    #lexi = update_lexi_payouts(lexi,offers) # we don't have conversions in SMS in the moment 
    lexi['Payout Revenue'] = lexi['Revenue']
    lexi['Payout eCPM'] = lexi['eCPM']
    lexi.sort_values(by=['Date'],inplace=True)
    #lexi['EMA Revenue'] = lexi.groupby(['SC_DP&Pub','Affiliate ID','Hitpath Offer ID','Super Segment'])['Payout Revenue'].transform(lambda x: x.ewm(alpha=0.4).mean())
    #lexi['EMA Clicks'] = lexi.groupby(['SC_DP&Pub','Affiliate ID','Hitpath Offer ID','Super Segment'])['Clicks'].transform(lambda x: x.ewm(alpha=0.4).mean())
    
#     cw = infrastructure.get_lanina()
#     lexi = update_creative_types(lexi,cw)
    
    #lexi = update_lexi_history(lexi)
    
    lexi = lexi[~lexi.duplicated(subset=['SC_DP&Pub','Affiliate ID','Hitpath Offer ID','Date','Segments','Revenue','Delivered','Send Strategy'])]
    lexi.sort_values(['Date','SC_DP&Pub','Affiliate ID'], inplace=True)
    
    print('Lexi Loading Complete!')    
    return lexi,lexi.copy(),lexi_raw

def load_el_nino():
#     class TimeoutError(Exception):
#         pass

#     def timeout_handler(signum, frame):
#         raise TimeoutError('Timed Out')

#     timeout_period = timeout_seconds

#     signal.signal(signal.SIGALRM, timeout_handler)
#     signal.alarm(timeout_period)

#     try:
    #cw = infrastructure.get_lanina()
    cw = infrastructure.get_lanina()
    cw.rename(columns={'Reporting Content ID':'Content ID'},inplace=True)
#     except TimeoutError as e:
#         creation_time = os.path.getmtime("ORIGINAL - La Nina - Content Schedule and Database - Content Warehouse.csv")
#         creation_time_formatted = datetime.datetime.fromtimestamp(creation_time).strftime("%m/%d/%y")

#         print(f"La Nina taking too long to load, defaulting to local copy from {creation_time_formatted}: ")
#         contentdf = pd.read_csv("ORIGINAL - La Nina - Content Schedule and Database - Content Warehouse.csv",dtype={'Reporting Content ID':str})
#         contentdf['Reporting Content ID']=contentdf['Reporting Content ID'].str[:7]
#         contentdf['Content Pitch']=np.where(contentdf['Type (Pitch)'].isna(),contentdf['Pitch Resolution'],contentdf['Type (Pitch)'])
#         cw = contentdf

#     finally:
#         signal.alarm(0)
        
    #cw['Seasonal/Holiday Date Restrictions'] = cw['Seasonal/Holiday Date Restrictions'].str.replace('11/31','11/30')
    cw = cw[~cw.duplicated(subset=['Content ID'])]
    return cw

def load_el_nino_timeout(timeout_seconds=15):
    class TimeoutError(Exception):
        pass

    def timeout_handler(signum, frame):
        raise TimeoutError('Timed Out')

    timeout_period = timeout_seconds

    signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(timeout_period)

    try:
        cw = infrastructure.get_lanina()
    except TimeoutError as e:
        creation_time = os.path.getmtime("ORIGINAL - La Nina - Content Schedule and Database - Content Warehouse.csv")
        creation_time_formatted = datetime.datetime.fromtimestamp(creation_time).strftime("%m/%d/%y")

        print(f"La Nina taking too long to load, defaulting to local copy from {creation_time_formatted}: ")
        contentdf = pd.read_csv("ORIGINAL - La Nina - Content Schedule and Database - Content Warehouse.csv",dtype={'Reporting Content ID':str})
        #contentdf['Reporting Content ID']=contentdf['Reporting Content ID'].str[:7]
        #contentdf['Content Pitch']=np.where(contentdf['Type (Pitch)'].isna(),contentdf['Pitch Resolution'],contentdf['Type (Pitch)'])
        cw = contentdf

    finally:
        signal.alarm(0)
        
    #cw['Seasonal/Holiday Date Restrictions'] = cw['Seasonal/Holiday Date Restrictions'].str.replace('11/31','11/30')
    cw = cw[~cw.duplicated(subset=['Content ID'])]
    
#     duplicated_columns = cw.columns[cw.columns.duplicated()]

#     cw = cw.drop(columns=[duplicated_columns[1]])

    return cw


    

def update_lexi_history(lexi):
    if os.environ.get('DOCKER_RUNNING'):
        cobra_auth_path='host_data/' + filepath.service_account_location.split('/')[-1]
    else:
        cobra_auth_path=filepath.service_account_location
        
    viper_settings_url = 'https://docs.google.com/spreadsheets/d/104t7KyOxfQvZH318MKxT0m7hMRRPvDMFVLIMd2YUVco/edit#gid=1354838607'
#     if os.environ.get('DOCKER_RUNNING'):
#         auth_path='host_data/' + filepath.service_account_location.split('/')[-1]
#     else:
#         auth_path=filepath.service_account_location
#     gc = pygsheets.authorize(service_account_file=cobra_auth_path)
    gc = infrastructure.get_gc()

#     gc = pygsheets.authorize(service_account_file=auth_path)
    viper_settings_sheet = gc.open_by_url(viper_settings_url)
    aff_settings_frame = viper_settings_sheet.worksheet_by_title('Affiliate Settings').get_as_df()
        
        
    
    aff_settings_frame.replace('',np.nan,inplace=True)
    history = aff_settings_frame[~aff_settings_frame['Previous Affiliate IDs'].isnull()]
    history['Previous Affiliate IDs'] = history['Previous Affiliate IDs'].astype(int)
    history[['Affiliate', 'Previous Affiliate IDs']] = history[['Affiliate', 'Previous Affiliate IDs']].astype(str)

    mapping = history.set_index('Previous Affiliate IDs')['Affiliate'].to_dict()
    for original_id, new_id in mapping.items():
        new_df = lexi[lexi['Affiliate ID']==new_id]
        original_df = lexi[lexi['Affiliate ID']==original_id]
        history_rows = original_df[original_df['Date']<new_df['Date'].min()].copy()
        history_rows['Affiliate ID'] = new_id
        lexi = pd.concat([lexi,history_rows],ignore_index=True)
    return lexi

def get_htm_el_nino():
    
#     if os.environ.get('DOCKER_RUNNING'):
#         cobra_auth_path='host_data/' + filepath.service_account_location.split('/')[-1]
#     else:
#         cobra_auth_path=filepath.service_account_location

#     gc = pygsheets.authorize(service_account_file=cobra_auth_path)
    gc = infrastructure.get_gc()
    
    url = 'https://docs.google.com/spreadsheets/d/1JLkc0Vb9UykY-X33SOdBo8Xy9qaZHlxtL3rKIAAZAdE/edit#gid=0'
    html_df = gc.open_by_url(url).worksheet('title','HTML Storage').get_as_df()

    return html_df
    

    
def get_lexi():
    if os.environ.get('DOCKER_RUNNING'):
        lexi_download_path='host_downloads/'
    else:
        lexi_download_path=filepath.downloadpath

    lexi_files = [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f))& ('SS_LC_merged_data' in f)]
    lexi_paths = ["{}/{}".format(lexi_download_path,lexi_version) for lexi_version in lexi_files]
    times = [os.path.getmtime(path) for path in lexi_paths]
    times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
    most_recent_lexi = lexi_files[times.index(max(times))]
    lexi_path = "{}/{}".format(lexi_download_path,most_recent_lexi)

    print('Loading Lexi Data...')
    lexi = pd.read_csv(lexi_path)
    lexi = infrastructure.transform_sms_df(lexi)
    lexi['eCPM'] = lexi['Revenue']*1000 / lexi['Delivered']
    lexi['CTR'] = lexi['Clicks']*1000 / lexi['Delivered']
#     lexi_clean = lexi.copy()
    print('Lexi Loading Complete!')
    
    offers = get_offers() #for proper merging
    lexi = pd.merge(lexi,offers[['Day Restrictions','Payout Type','Scheduling Name','Status','Paused Pubs']],left_on='Hitpath Offer ID', right_on=offers.index, how='left')
    
    lexi['Affiliate ID'] = lexi['Affiliate ID'].astype(float).astype(int).astype(str) #sometimes Lexi loads with varying dtypes

    #Deal with Duplicate Affiliate IDs for different DP&Pubs
    #lexi.loc[ lexi['DP&Pub']=='SC.FHA_460398', 'Affiliate ID'] = '160398'
    #lexi.loc[ lexi['DP&Pub']=='SC.RF_460398', 'Affiliate ID'] = '260398'
    #lexi.loc[ lexi['DP&Pub']=='LPG.FHA_461128', 'Affiliate ID'] = '161128'
    #lexi.loc[ lexi['DP&Pub']=='LPG.RF_461128', 'Affiliate ID'] = '261128'
    
    return lexi,lexi



def get_content_report():
    
    if os.environ.get('DOCKER_SCRIPT_RUNNING'):
        from google.cloud import storage

        print('Checking viper-bucket')
        # viper-bucket/Lexi 3.0 - 2023-12-04T142046.393.xlsx
        bucket_name = "viper-bucket"
        auth_file_name = 'Custom Content Scheduling Report_5786_Optimizations.xlsx'
        # folder_path = 'Authentication_and_Data/'
        local_auth_directory = '/tmp/' 

        # os.makedirs(local_auth_directory, exist_ok=True)

        cc_file_path = local_auth_directory + auth_file_name

        storage_client = storage.Client()

        bucket = storage_client.bucket(bucket_name)
        blob = bucket.blob(auth_file_name)
#         print('blob is',blob)

        blob.download_to_filename(cc_file_path)
    
        content_report = pd.ExcelFile(cc_file_path)
    
    elif os.environ.get('DOCKER_RUNNING'):
        cc_download_path='host_downloads/'
        
        cc_files = [f for f in listdir(cc_download_path) if isfile(join(cc_download_path, f)) & ('Custom Content Scheduling Report_5786_Optimizations' in f)]
        cc_paths = ["{}/{}".format(cc_download_path,cc_version) for cc_version in cc_files]
        times = [os.path.getctime(path) for path in cc_paths]
        # times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
        most_recent_cc = cc_files[times.index(max(times))]
        most_recent_cc_file = "{}/{}".format(cc_download_path,most_recent_cc)

        content_report = pd.ExcelFile(most_recent_cc_file)
    else:
        cc_download_path=filepath.downloadpath

        cc_files = [f for f in listdir(cc_download_path) if isfile(join(cc_download_path, f)) & ('Custom Content Scheduling Report_5786_Optimizations' in f)]
        cc_paths = ["{}/{}".format(cc_download_path,cc_version) for cc_version in cc_files]
        times = [os.path.getctime(path) for path in cc_paths]
        # times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
        most_recent_cc = cc_files[times.index(max(times))]
        most_recent_cc_file = "{}/{}".format(cc_download_path,most_recent_cc)

        content_report = pd.ExcelFile(most_recent_cc_file)
        
    return content_report

# def get_redrop_report():
#     if os.environ.get('DOCKER_RUNNING'):
#         redrop_download_path='host_downloads/'
#     else:
#         redrop_download_path=filepath.downloadpath
        
#     redrop_files = [f for f in listdir(redrop_download_path) if isfile(join(redrop_download_path, f)) & ('Detailed Version - Offer Redrop Report (' in f)]
#     redrop_versions = [int(num.split('(')[1].split(')')[0]) for num in redrop_files]

#     most_recent_redrop = max(redrop_versions)

#     redrop_report = pd.ExcelFile(redrop_download_path+'Detailed Version - Offer Redrop Report ({}).xlsx'.format(most_recent_redrop))
#     return redrop_report

def get_test_report():
    if os.environ.get('DOCKER_RUNNING'):
        download_path='host_downloads/'
    else:
        download_path=filepath.downloadpath
    test_files = [f for f in listdir(download_path) if isfile(join(filepath.downloadpath, f)) & ('Redrop - ' in f)]
    test_paths = ["{}/{}".format(download_path,test_version) for test_version in test_files]
    times = [os.path.getmtime(path) for path in test_paths]
    most_recent_test = test_files[times.index(max(times))]
    test_path = "{}/{}".format(download_path,most_recent_test)
    test_report = pd.read_excel(pd.ExcelFile(test_path), sheet_name= 'Offer Recommendations')
    return test_report
 
def get_swap_report(name,offers):
    if os.environ.get('DOCKER_RUNNING'):
        swap_download_path='host_downloads/'
    else:
        swap_download_path=filepath.downloadpath

    swap_files = [f for f in listdir(swap_download_path) if isfile(join(swap_download_path, f)) & ('swap_report' in f)]
    swap_paths = ["{}/{}".format(swap_download_path,swap_version) for swap_version in swap_files]
    times = [os.path.getmtime(path) for path in swap_paths]
    most_recent_swap = swap_files[times.index(max(times))]
    
    swap_path = "{}/{}".format(swap_download_path,most_recent_swap)
    
    swap_report = pd.read_excel(pd.ExcelFile(swap_path))
    if len(swap_report) > 0: 
        swap_report['Affiliate ID'] = swap_report['PUBID'].astype(str).str.split('.',expand=True)[0]  
        swap_report['Hitpath Offer ID'] = swap_report['Offer'].str.split('-| ',expand=True)[0]  
        swap_report['Payout Type'] = np.where(swap_report['Hitpath Offer ID'].notnull(),
                                          swap_report['Hitpath Offer ID'].map(offers['Payout Type']),'')

        swap_report = swap_report[swap_report['Responsible DMA']==name]
    
#     swap_report['Payout Type'] = swap_report['Hitpath Offer ID'].apply(lambda hit: offers.loc[hit]['Payout Type'] if hit in offers len(offers[offers['Hitpath Offer ID']==int(hit)].values)>0 else'')
    
    return swap_report

def refresh_cobra(days_ahead=60,secure=False):
#     global cobra
#     global cobra_clean
#     global cobra_sheet
#     global cobra_sheet_df
    
    print('Loading Cobra...')
    if secure==False:
        cobra= infrastructure.get_upcoming_mamba(days_ahead=days_ahead)
    else:
        cobra=infrastructure.get_cobra_secure() # we don't have get_cobra_secure() in SMS yet 
    
    cobra['Affiliate ID']=cobra['Dataset'].str.split('_').str[2]
    #cobra[['Akshad Notes:','real seg']] = cobra[['Akshad Notes:','real seg']].fillna("")
    cobra['Drop Number'] = cobra['Drop'].str.split(expand=True)[1].astype(int)
    cobra['Day of Week'] = cobra['Date'].dt.dayofweek
#     cobra.dropna(subset=['Hitpath Offer ID'], inplace=True)
    #cobra['Campaign ID'] = cobra['Campaign ID'].replace(np.nan, '')
    #cobra['Hitpath Offer ID'] = cobra['Campaign ID'].apply(lambda x: x.split('-')[0][:5])
    #cobra['Campaign ID'] = cobra['Campaign ID'].replace(r'^\s*$',np.nan,regex=True)
#     cobra = cobra.astype({"Hitpath Offer ID": int})
    #cobra['Hitpath Offer ID']=pd.to_numeric(cobra['Hitpath Offer ID'],errors='coerce') #sometimes Hitpath Offer ID is a string
    cobra.rename(columns={"Segment ": "Segment"},inplace=True)
    
    cobra_clean = cobra.copy()

    gc = infrastructure.get_gc()
    
    cobra_sheet = gc.open_by_url('https://docs.google.com/spreadsheets/d/12vqSDueybprNphtsw7gXR5vmgcPG6_5ZNcnWzNpiasY/edit#gid=534096291').worksheet('title','New Mamba')

    # open worksheet - cobra
    # schedule_wks  = cobra_sheet[1]
    cobra_sheet_df = cobra_sheet.get_as_df()

    print('Cobra Loaded!')
    #reformat to resemble cobra format
    cobra_sheet_df.columns = cobra_sheet_df.iloc[0]

    s = cobra_sheet_df.columns.to_series()
    s.iloc[0] = 'A'
    s.iloc[1] = 'B'
    cobra_sheet_df.columns = s
    cobra_sheet_df.index = np.arange(2, len(cobra_sheet_df) + 2)
    
    
    last_cobra_date_string = cobra_sheet_df.columns[-1]
    last_cobra_date = datetime.datetime.strptime(last_cobra_date_string, '%m/%d/%Y').date()

    start_date = last_cobra_date + timedelta(days=1)
    end_date = date.today() + timedelta(days=days_ahead)

    new_date_list = list(pd.date_range(start_date,end_date,freq='d'))
    new_date_list_columns = [date_obj.strftime('%-m/%-d/%Y') for date_obj in new_date_list]


    all_date_columns = list(cobra_sheet_df.columns) + new_date_list_columns
    cobra_sheet_df = cobra_sheet_df.reindex(columns=all_date_columns)
    cobra_sheet_df.loc[:2] = all_date_columns

    
       
    return cobra, cobra_clean, cobra_sheet, cobra_sheet_df


def get_affiliate_log_template():
    if os.environ.get('DOCKER_RUNNING'):
        service_account_path='host_data/' + filepath.service_account_location.split('/')[-1]
    else:
        service_account_path=filepath.service_account_location

#     gc = pygsheets.authorize(service_account_file=service_account_path)
    gc = infrastructure.get_gc()
    
    sh = gc.open_by_url('https://docs.google.com/spreadsheets/d/1oKod0EKYfZmrS9MUtmftmEcv6WPESP5-qMJ6DBuJ4kU/edit#gid=0') # sms viper log

    template_columns = sh.worksheets(sheet_property='title',value='Template')[0][1]
    template_columns = [tc for tc in template_columns if tc!='']
    template_log = pd.DataFrame(columns=template_columns)
    return template_log

def update_creative_types(lexi,cw):
    #if a reporting content id is placed in the creative type column replace it with the content id from La Nina
    df = lexi.copy()
    df_subset = df[df['Creative Type'].astype(str).str.match(r'^1\d{6}$')][['Creative Type']].rename({'Creative Type':'Reporting Content ID'},axis=1)
    df_subset = pd.merge(df_subset, cw[['Reporting Content ID','Content ID']], how='left', on='Reporting Content ID').drop_duplicates()
    creative_mapping = df_subset.set_index('Reporting Content ID')['Content ID'].to_dict()
    df['Creative Type'].replace(creative_mapping,inplace=True)
    
    return df

def update_lexi_payouts(lexi,offers,new_payout_entries=[]):

    #**** Build Direct Agents Payouts ****

    direct_agents_url = 'https://docs.google.com/spreadsheets/d/1iUxuhBWag0Pamg-GVLpfiJo-ENc2AQzdw3uYhAXVEcQ/edit#gid=1337786795'
    if os.environ.get('DOCKER_RUNNING'):
        service_account_path='host_data/' + filepath.service_account_location.split('/')[-1]
    else:
        service_account_path=filepath.service_account_location
        
#     gc = pygsheets.authorize(service_account_file=service_account_path)
    gc = infrastructure.get_gc()

    da_sheet = gc.open_by_url(direct_agents_url)
    direct_agent_names = [ws.title for ws in da_sheet.worksheets() if 
                         ((ws.title!='TEMPLATE') & ~('copy' in ws.title.lower())
                          &~('formatted da data' in ws.title.lower()))]

    col_names = ["PUBID","Status","DATE - STATUS CHANGE","Old Rate","New Rate","1","2","3","4","5","6","7","8","9","10","11","12","13","14"]

    df_all_direct_agents = pd.DataFrame(columns=col_names) #create empty df

    for da_name in direct_agent_names:
        da_sheet = gc.open_by_url(direct_agents_url).worksheet('title',da_name)
        da_offers = list(lexi[lexi['Scheduling Name'].str.contains(da_name,na=False)]['Hitpath Offer ID'].unique())
        da_frame = da_sheet.get_as_df(start=(2,1))
        da_frame.drop(da_frame.columns[5:],axis=1,inplace=True)
        da_frame.rename({'NEW RATE':'New Rate','OLD RATE':'Old Rate','STATUS':'Status'},axis=1,inplace=True)
        da_frame = da_frame.set_index('PUBID')
        da_frame['Offers'] = [da_offers] * len(da_frame)
        da_frame = da_frame.explode('Offers')
        da_frame = da_frame[(~da_frame['New Rate'].isnull()) & ~(da_frame['New Rate']=='')]
        da_frame = da_frame.reset_index()
        df_all_direct_agents = pd.concat([df_all_direct_agents, da_frame])

    df_all_direct_agents['PUBID'] = df_all_direct_agents['PUBID'].astype(str)
    dups = df_all_direct_agents[df_all_direct_agents['PUBID']=='460398']
    df_all_direct_agents['PUBID'].replace('460398','160398',inplace=True)
    dups['PUBID'].replace('460398','260398',inplace=True)
    df_all_direct_agents = pd.concat([dups,df_all_direct_agents])

    dups = df_all_direct_agents[df_all_direct_agents['PUBID']=='461128']
    df_all_direct_agents['PUBID'].replace('461128','161128',inplace=True)
    dups['PUBID'].replace('461128','261128',inplace=True)
    df_all_direct_agents = pd.concat([dups,df_all_direct_agents])

    lexi = pd.merge(lexi, df_all_direct_agents[['PUBID','Offers','New Rate','Old Rate']], left_on=['Affiliate ID','Hitpath Offer ID'], right_on=['PUBID','Offers'], how='left', suffixes=['_'+da_name+'_x','_'+da_name+'_y'])

    lexi['Old Rate'].fillna(lexi['New Rate'],inplace=True)

    #**** Build All Other Offers Payouts ****  

    affs = lexi['Affiliate ID'].unique()
    hits = lexi['Hitpath Offer ID'].unique()

    lexi['Payout'] = np.where(lexi['Payout Type'].isin(['CPA','CPC','CPL']), lexi['Revenue']/lexi['Conversions'], np.nan)
    lexi['Payout'] = lexi['Payout'].fillna(1)

    multi_index = pd.MultiIndex.from_product([affs, hits], names=['Affiliate ID','Hitpath Offer ID'])
    payout_frame = pd.DataFrame(0, columns=['Payout'],index=multi_index)

    def payout_split(payout_string):
        payout_string = str(payout_string).replace('for', ' ').replace(':','').replace(',','')
        return [x.split() for x in str(payout_string).split("\n")]

    offers['Payout Split'] = offers['Payout'].apply(lambda x: payout_split(x))

    offers = offers.explode('Payout Split')

    offer_payouts = offers[(offers['Payout Split'].str.len()==2) 
        &
        ( 
        (  offers['Payout Split'].str[0].str.isdigit()) 
        & (offers['Payout Split'].str[1].str.startswith('$', na=False))
        )
        |
        ( (offers['Payout Split'].str[1].str.isdigit()) 
        & (offers['Payout Split'].str[0].str.startswith('$', na=False))
        )
        ]['Payout Split']

    offer_payouts = pd.DataFrame(offer_payouts)
    offer_payouts['Affiliate ID'] = offer_payouts['Payout Split'].apply(lambda x: x[0] if len(x[0])==6 else x[1])
    offer_payouts['New Payout'] = offer_payouts['Payout Split'].apply(lambda x: x[1] if len(x[0])==6 else x[0])
    offer_payouts['New Payout'] = offer_payouts['New Payout'].str.replace('$','', regex=False)
    offer_payouts['New Payout'] = pd.to_numeric(offer_payouts['New Payout'],errors='coerce')
    offer_payouts = offer_payouts.reset_index()


    payout_frame = payout_frame.reset_index()
    payout_frame.drop('Payout',axis=1,inplace=True)

    payout_frame = pd.merge(payout_frame,offer_payouts,left_on=['Affiliate ID','Hitpath Offer ID'],right_on=['Affiliate ID','Hitpath Offer ID'],how='left')

    simple_payouts = offers[~(offers['Scheduling Name'].str.contains('Direct Agents',na=False))&(offers['Payout Split'].str.len()==1)&(offers['Payout'].str.lower()!='variable')].rename({'Payout':'Simple Payout'},axis=1)[['Simple Payout']]
    simple_payouts['Simple Payout'] = pd.to_numeric(simple_payouts['Simple Payout'],errors='coerce')

    payout_frame = pd.merge(payout_frame, simple_payouts, how='left', on='Hitpath Offer ID')
    payout_frame['New Payout'].fillna(0,inplace=True)
    payout_frame['Simple Payout'].fillna(0,inplace=True)
    payout_frame['New Payout'] = payout_frame[["New Payout", "Simple Payout"]].values.max(axis=1)
    payout_frame.drop('Simple Payout',axis=1,inplace=True)


    lexi = pd.merge(lexi, payout_frame, left_on=['Affiliate ID','Hitpath Offer ID'], right_on=['Affiliate ID','Hitpath Offer ID'], how='left')  

    lexi['New Rate'] = lexi['New Rate'].str.replace('$','', regex=False)
    lexi['New Rate'].fillna(0,inplace=True)
    lexi['New Rate'] = pd.to_numeric(lexi['New Rate'],errors='coerce')
    lexi['New Payout'] = pd.to_numeric(lexi['New Payout'],errors='coerce')
    lexi['New Payout'].fillna(0,inplace=True)
    lexi['Updated Payout'] = lexi[["New Rate", "New Payout"]].values.max(axis=1)

    lexi['Old Rate'] = lexi['Old Rate'].str.replace('$','', regex=False)
    lexi['Old Rate'] = pd.to_numeric(lexi['Old Rate'],errors='coerce')
    lexi['Old Rate'].fillna(1,inplace=True)
    lexi['Outdated Payout'] = lexi[["Old Rate", "Payout"]].values.max(axis=1)

    new_payout_entries=[]

    if len(new_payout_entries)==3:

        new_payout_affiliate = new_payout_entries[0]
        new_payout_hitpath = new_payout_entries[1]
        new_payout = new_payout_entries[2]

        to_update = (lexi['Affiliate ID'] == new_payout_affiliate) & (lexi['Hitpath Offer ID'] == new_payout_hitpath)

        lexi.loc[ to_update, 'Updated Payout'] = new_payout


    lexi['Payout eCPM'] = np.where((~lexi['Updated Payout'].isnull()&~lexi['Payout'].isnull()) , lexi['eCPM']*lexi['Updated Payout']/lexi['Outdated Payout'], lexi['eCPM'])
    lexi['Payout Revenue'] = np.where((~lexi['Updated Payout'].isnull()&~lexi['Payout'].isnull()) , lexi['Revenue']*lexi['Updated Payout']/lexi['Outdated Payout'], lexi['Revenue'])
    
    # Manual Corrections for bad data
    lexi.loc[lexi['Payout eCPM']==0,'Payout eCPM'] = lexi.loc[lexi['Payout eCPM']==0, 'eCPM']
    lexi.loc[lexi['Payout Revenue']==0,'Payout Revenue'] = lexi.loc[lexi['Payout Revenue']==0, 'Revenue']
    lexi.loc[~lexi['Payout Type'].isin(['CPA','CPC','CPL']),'Payout Revenue'] = lexi.loc[~lexi['Payout Type'].isin(['CPA','CPC','CPL']), 'Revenue']
    lexi.loc[~lexi['Payout Type'].isin(['CPA','CPC','CPL']),'Payout eCPM'] = lexi.loc[~lexi['Payout Type'].isin(['CPA','CPC','CPL']), 'eCPM']

    lexi['Abs Rev Diff'] = abs((lexi['Payout Revenue'] - lexi['Revenue'])/lexi['Revenue'])
    lexi.loc[lexi['Abs Rev Diff']>0.5,'Payout Revenue'] = lexi.loc[lexi['Abs Rev Diff']>0.5, 'Revenue'] 
    
    lexi.drop(columns = ['PUBID','Offers','New Payout','Old Rate', 'New Rate', 'Payout Split', 'Abs Rev Diff','Outdated Payout'], inplace=True)
    
    return lexi


def complete_lexi_payouts(lexi, new_payout_affiliate, new_payout_hitpath, new_payout):

    to_update = (lexi['Affiliate ID'] == new_payout_affiliate) & (lexi['Hitpath Offer ID'] == new_payout_hitpath)

    lexi.loc[ to_update, 'Updated Payout'] = new_payout
    
    
    lexi['eCPM'] = np.where((~lexi['Updated Payout'].isnull()&~lexi['Payout'].isnull()) , lexi['eCPM']*lexi['Updated Payout']/lexi['Payout'], lexi['eCPM'])

    return lexi

def upload_lexi_to_bucket():    
    storage_client = storage.Client()

    bucket_name = "viper-bucket"
    bucket = storage_client.bucket(bucket_name)

    blob = bucket.blob('Lexi.xlsx')
    if blob.exists():
        blob.delete()
        print(f'Old Lexi Removed from {bucket_name}.')

    lexi_download_path=filepath.downloadpath

    lexi_files = [f for f in listdir(lexi_download_path) if isfile(join(lexi_download_path, f)) 
                  & ('Lexi 3.0' in f) & ~("Year" in f)]
    lexi_paths = ["{}/{}".format(lexi_download_path,lexi_version) for lexi_version in lexi_files]
    times = [os.path.getmtime(path) for path in lexi_paths]
    # times_formatted = [time.strftime('%m/%d/%Y %H:%M:%S', time.gmtime(time_float)) for time_float in times]
    most_recent_lexi = lexi_files[times.index(max(times))]
    lexi_path = "{}/{}".format(lexi_download_path,most_recent_lexi)
    
    new_blob = bucket.blob('Lexi.xlsx')
    new_blob.upload_from_filename(lexi_path, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    print(f'New Lexi Uploaded to {bucket_name}.')

# def log_aff_to_df(aff, sh, all_sheet_titles):
#     print('in viper main')
#     if __name__ == 'viper_main':
#         if aff in all_sheet_titles:
#             aff_df = sh.worksheets(sheet_property='title',value=aff)[0].get_as_df()
#             aff_df['Affiliate ID'] = aff #clean up for CPAM
#         else:
#             aff_df = sh.worksheets(sheet_property='title',value='Template')[0].get_as_df()
#             aff_df['Affiliate ID'] = aff #clean up for CPAM

#     return aff_df

# def square(x):
#     return x ** 2