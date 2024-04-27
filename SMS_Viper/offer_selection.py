import datetime
from datetime import timedelta
from datetime import date
import numpy as np
import pandas as pd
import sys
import find_viper_settings
import scheduling
import infrastructure
import viper_main
import os
import filepath
import time
import pygsheets


class inputs:
    def __init__(self, fvs=None, lexi=None, cobra=None, offers=None, emit=None, cs=None, today=date.today()):
#         self.lock = threading.Lock()
        self.today = today

        self.fvs = fvs
        self.cobra = cobra
        self.emit = emit
        self.offers = offers
        self.lexi = lexi
        self.lexi_90 = self.lexi[(self.lexi['Date'].dt.date >= (today + timedelta(days=-90)))]
#         self.functions_dict = functions_dict
        self.cs = cs
        self.mirrors,self.true_mirrors = fvs.find_mirrors()
        
#         url = 'https://docs.google.com/spreadsheets/d/104t7KyOxfQvZH318MKxT0m7hMRRPvDMFVLIMd2YUVco/edit#gid=1354838607'
#         sheets = Sheets.from_files(filepath.gsheets)
#         self.auto_scheduling_sheet = sheets.get(url)
        
        
        viper_settings_url = 'https://docs.google.com/spreadsheets/d/104t7KyOxfQvZH318MKxT0m7hMRRPvDMFVLIMd2YUVco/edit#gid=1354838607'
#         if os.environ.get('DOCKER_RUNNING'):
#             auth_path='host_data/' + filepath.service_account_location.split('/')[-1]
#         else:
#             auth_path=filepath.service_account_location

#         gc = pygsheets.authorize(service_account_file=auth_path)
#         self.gc, self.auth_path = infrastructure.get_gc()

#         self.viper_settings_sheet = gc.open_by_url(viper_settings_url)
        self.viper_settings_sheet = fvs.viper_settings_sheet
        
        
        
        self.redrop_ot_report = self.generate_redrops_and_ots('all')
        
        
        

    def find_budgeted_offers(self,sc_dppub_affiliate):
        cobra = self.cobra
        
#         budget_frame = find_viper_settings.get_budget_frame()
        fvs = self.fvs
        budget_frame = fvs.get_budget_frame()
        
        grouped_budget_frame = budget_frame.groupby(['PUB ID','Scheduling Name ']).agg({'Production Segment':'sum'}).reset_index()
        grouped_budget_frame['Scheduling Name '] = grouped_budget_frame['Scheduling Name '].astype(str)
        grouped_budget_frame['Hitpath Offer ID'] = grouped_budget_frame['Scheduling Name '].str.split('-').str[0].astype(int)
        grouped_budget_frame['PUB ID'] = grouped_budget_frame['PUB ID'].astype(str)
        budgeted_offers_affiliate = grouped_budget_frame[grouped_budget_frame['PUB ID']==sc_dppub_affiliate]
        budgeted_offers_list = budgeted_offers_affiliate['Hitpath Offer ID'].tolist()
        
    #     budgeted_offers_list = [x for x in budgeted_offers_list if x not in exclude_offers]
    #     budgeted_offers_list = scheduling.return_unpaused_offers(sc_dppub_affiliate, budgeted_offers_list)

        budget_title_month = fvs.get_budget_title_month()
        
        budgeted_offers_list_return = []
        for hitpath in budgeted_offers_list:
            budgeted_count = len(cobra[(cobra['shortcode_DP.SV']==sc_dppub) & (cobra['Campaign ID'].str.contains(str(hitpath))) & 
                                        (cobra['Date'].dt.month_name()==budget_title_month)])

            budget_entry = grouped_budget_frame[(grouped_budget_frame['PUB ID']==sc_dppub_affiliate) & (grouped_budget_frame['Hitpath Offer ID']==hitpath)]['Production Segment'].values[0]

            try:
                allocated_count = int(budget_entry)

            except:
                allocated_count = 0
                for character in budget_entry: #will break with double digit entry
                    if character.isdigit():
                        allocated_count = int(character)
            
            if budgeted_count < allocated_count:
                budgeted_offers_list_return.append(hitpath)
#         print(budgeted_offers_list_return)
        return budgeted_offers_list_return

    def find_top_p_offers(self,sc_dppub_affiliate,primary_metric='eCPM',ema=True,top_number=10):
        lexi = self.lexi
        fvs = self.fvs
        
        #find the average eCPM (or other chosen primary metric) for the A120 segment of the account being scheduled in the last 30 days
        #this average is our "benchmark"
        #any live/budgeted offers that have had a drop in the last month exceed the benchmark found will be added to the "top offers"
        #** comment it as we don't need it for now for sms 
        # budgeted_offers_to_schedule = self.find_budgeted_offers(sc_dppub_affiliate)
        
#         recent_days = lexi['Date'].max() - timedelta(days=30)
#         affiliate_recent =  lexi[(lexi['SC_DP&Pub']==sc_dppub) & (lexi['Date']>=recent_days) ]
#         affiliate_recent_production = affiliate_recent[((affiliate_recent['Send Strategy']=='P')
#                                                       |(affiliate_recent['Segment'].str.contains('A120',na=False)))
#                                                       & (affiliate_recent['Status']=='Live') ]
#         benchmark = affiliate_recent_production['Revenue'].sum()*1000 / affiliate_recent_production['Delivered'].sum()

        
        df = lexi[lexi['SC_DP&Pub']==sc_dppub_affiliate]
        df = df[df['Date'].dt.date>=date.today()+timedelta(days=-365)]
        
        lexi_running = fvs.lexi_running.copy()
        
        #eCPM

#         df['Account P eCPM'] = df[['Date','SC_DP&Pub']].apply(lambda x: fvs.find_ecpm_from_date(x['SC_DP&Pub'], 
#                                                          x['Date'], days_back=30, ema=True, metric='Revenue'),axis=1)
        
        df = pd.merge(df, lexi_running[['30 Day EMA eCPM']], how='left', on=['SC_DP&Pub','Date'])
        df.rename(columns={'30 Day EMA eCPM':'Account P eCPM'},inplace=True)
        

        df['EMA eCPM'] = df['EMA Revenue']*1000 / df['Delivered']
        df['EMA eCPM Ratio'] = df['EMA eCPM'] / df['Account P eCPM']
        
        #CTR50

#         df['Account P CTR'] = df[['Date','SC_DP&Pub']].apply(lambda x: fvs.find_ecpm_from_date(x['SC_DP&Pub'], 
#                                                              x['Date'], days_back=30, ema=True, metric='Clicks'),axis=1)
        df = pd.merge(df, lexi_running[['30 Day EMA CTR']], how='left', on=['SC_DP&Pub','Date'])
        df.rename(columns={'30 Day EMA CTR':'Account P CTR'},inplace=True)


        df['EMA CTR'] = df['EMA Clicks']*1000 / df['Delivered']
        df['EMA CTR50'] = ((df['EMA eCPM']/df['Account P eCPM']) + (df['EMA CTR']/df['Account P CTR']))/2
        df['Payout Revenue'] = df['Revenue']

        #Offers
        affiliate_recent = df[df['Date']>=lexi['Date'].max() - timedelta(days=30)]
        
        aff_offers = affiliate_recent[(affiliate_recent['Status'].isin(['Live','Paused - Budget']))  &
                                                 (~affiliate_recent.isin(['CPM']).any(axis=1)) &
                    (affiliate_recent['Send Strategy'].isin(['P','CT','E','OT','PT','RT','IT','MA']))]
        # since we are holding payout_split in viper_main, payout Revenue = revenue 


        aff_offers = aff_offers.groupby(['Date','Hitpath Offer ID']).agg({'Payout Revenue':'sum','Delivered':'sum','CTR':'mean','Account P eCPM':'first','Account P CTR':'first'})
        
        
        aff_offers['Payout eCPM'] = aff_offers['Payout Revenue']*1000 / aff_offers['Delivered']
        aff_offers['Payout CTR50'] = ((aff_offers['Payout eCPM']/aff_offers['Account P eCPM']) + 
                                          (aff_offers['CTR']/aff_offers['Account P CTR']))/2
        
        if primary_metric.lower()=='ctr50':
#             df['Account P CTR'] = df[['Date','SC_DP&Pub']].apply(lambda x: fvs.find_ecpm_from_date(x['SC_DP&Pub'], 
#                                                              x['Date'], days_back=30, ema=True, metric='Clicks'),axis=1)
#             df['EMA CTR'] = df['EMA Clicks']*1000 / df['Delivered']
#             df['EMA CTR50'] = ((df['EMA eCPM']/df['Account P eCPM']) + (df['EMA CTR']/df['Account P CTR']))/2
            
            best_offers_metric = 'EMA CTR50'
            all_offers_metric = 'Payout CTR50'

            benchmark = df['Account P CTR'].iloc[-1]
        else:
            best_offers_metric = 'EMA eCPM Ratio'
            all_offers_metric = 'Payout eCPM'
            
            benchmark = df['Account P eCPM'].iloc[-1]
            

        aff_offers.reset_index(inplace=True)

        offers_to_schedule = list(set(aff_offers[aff_offers[all_offers_metric]>=benchmark]['Hitpath Offer ID']))

        
        best_offers_df = df[(df['Status']=='Live')&
                            (df['Send Strategy'].isin(['P','CT','E','OT','PT','RT','IT','MA']))].groupby(['Hitpath Offer ID']).agg({best_offers_metric:'last'}) #lexi df should already be sorted

        best_offers = best_offers_df.sort_values(best_offers_metric,ascending=False).head(top_number).index.tolist()

        offers_to_schedule = list(set(offers_to_schedule + best_offers))

        
        #sort so that we schedule offers with more day restrictions earlier
        #secondary sort is by most recent eCPM
        #(manual offers will take priority above all)
        
        offers_to_schedule = sorted(offers_to_schedule, key = lambda hit : (len(fvs.find_days(hit)),-self.find_last_ecpm(sc_dppub_affiliate,hit)) )   
#         print(offers_to_schedule)
#         print(offers_to_schedule)
#         for o in offers_to_schedule:
#             print(o, self.find_last_ecpm(sc_dppub_affiliate,o))
#         print('budget',budgeted_offers_to_schedule)          
        return offers_to_schedule


    def get_redrop_offers(self,sc_dppub_affiliate):
        lexi = self.lexi
        emit = self.emit
        fvs = self.fvs
        
        print('Using Outdated Redrop Function!')

        redrop_report = viper_main.get_redrop_report()
#         fvs.set_lexi(lexi)
#         fvs.set_offers(offers)

        dp_pub = emit[emit['Revenue Pub ID']==int(sc_dppub_affiliate)]['DP&Pub'].mode().values[0]
        if dp_pub in redrop_report.sheet_names:
            redrops = pd.read_excel(redrop_report, dp_pub)
        else:
            redrops = pd.DataFrame({'Hitpath Offer ID' : []})

        fillers = pd.concat([redrops])

        requests_redrops_list = fvs.find_AM_requests(sc_dppub_affiliate,'redrop')

        filler_hitpaths = sorted(fillers['Hitpath Offer ID'].tolist(), key = lambda hit : len(fvs.find_days(hit)))
        filler_hitpaths = requests_redrops_list + filler_hitpaths #put AM requests at the front of redrop queue

        return filler_hitpaths

    def get_last_resort_offers(self,sc_dppub_affiliate):
        #Return all offers in the last 90 days from this account sorted by CTR50
        lexi_90 = self.lexi_90

        affiliate_extended =  lexi_90[(lexi_90['SC_DP&Pub']==sc_dppub_affiliate) ]

    #     ecpm_benchmark = affiliate_extended['Revenue'].sum()*1000 / affiliate_extended['Delivered'].sum()

        last_resort_offers = (affiliate_extended[(affiliate_extended['Payout Type']!='CPM') &
                          ((affiliate_extended['Send Strategy']=='P') | (affiliate_extended['Send Strategy']=='E') | (affiliate_extended['Send Strategy']=='OT'))  &
                          ((affiliate_extended['Status']=='Live') )
                         ]).sort_values(by='CTR50',ascending=False)['Hitpath Offer ID'].unique().tolist()

        return last_resort_offers     


    def find_top_mi_offers(self,sc_dppub_affiliate,segment='default'):
        # This function will return a ranked list of offers to mine based on Open Rate.  Note however that there will very likely be duplicates of the same offer in this list.  Viper will start from the first unpaused offer in the list and schedule it in the next mining slot available for that offer.  For example, if the next 7 mining slots are empty for an account and this function returns mining_offers = [5786,5786,5786,4032,5786,5786,8025] then that is the order in which the offers will be scheduled in the earliest availabe slot for that offer (taking day restrictions into account)
        lexi = self.lexi
        fvs = self.fvs

        recent_days = lexi['Date'].max() - timedelta(days=30)
        affiliate_recent =  lexi[(lexi['SC_DP&Pub']==sc_dppub) & (lexi['Date']>=recent_days)]

        production_pointer = 0
        mining_pointer = 0
        mining_offers = []

        affiliate_recent_mining = affiliate_recent[(affiliate_recent['Send Strategy']=='MI') & (affiliate_recent['Status']=='Live')].sort_values(by='Open Rate', ascending=False)   
        
        if segment!='default':
            affiliate_recent_mining = affiliate_recent_mining[(affiliate_recent_mining['Segment']==segment)]

        affiliate_recent_p_ctr = affiliate_recent[(affiliate_recent['Send Strategy']=='P') &
                         (affiliate_recent['Status']=='Live')
                        ].sort_values(by='Open Rate', ascending=False)


        for i in range(len(affiliate_recent_mining)+len(affiliate_recent_p_ctr)):
            source = np.random.choice(['MI', 'P'],p=[0.25, 0.75])

            if (source == 'MI') & (mining_pointer<len(affiliate_recent_mining)):
                offer = affiliate_recent_mining.iloc[mining_pointer]['Hitpath Offer ID']
#                 open_rate = affiliate_recent_mining.iloc[mining_pointer]['Open Rate']
#                 print(f'Offer {offer} with Open Rate {open_rate} from Mining')
                mining_offers.append(offer)
                mining_pointer+=1
            elif production_pointer<len(affiliate_recent_p_ctr):
                offer = affiliate_recent_p_ctr.iloc[production_pointer]['Hitpath Offer ID']
                mining_offers.append(offer)
                production_pointer+=1   

        budgeted_offers_list = fvs.get_budgeted_offers_list(sc_dppub_affiliate)
        return [x for x in mining_offers if x not in budgeted_offers_list]

    def find_last_ecpm(self,sc_dppub_affiliate,hitpath):
        lexi = self.lexi
        last_drop = lexi[(lexi['SC_DP&Pub']==sc_dppub_affiliate) & (lexi['Hitpath Offer ID']==hitpath)
                    & (lexi['Send Strategy'].isin(['P','CT','E','OT','PT','RT','IT','MA']))  ].iloc[-1]
    #     last_ecpm = (last_drop['Revenue'].sum()*1000) / last_drop['Delivered'].sum() 
        # we are holding payout_split in viper_main, payout eCPM = eCPM
        last_drop['Payout eCPM'] = last_drop['eCPM']
        last_ecpm = last_drop['Payout eCPM'] #Using updated payout eCPM here
        return last_ecpm   


    def find_top_advertiser_offers(self,sc_dppub_affiliate,advertiser):
        lexi = self.lexi
        offers = self.offers

        advertiser_offers = offers[(offers['Advertiser Name']==advertiser) & 
                                   ((offers['Status']=='Live') ) ].index.tolist()

        advertiser_offers = self.return_unpaused_offers(sc_dppub_affiliate, advertiser_offers)

        affiliate_advertiser_drops = lexi[(lexi['SC_DP&Pub']==sc_dppub_affiliate) & (lexi['Hitpath Offer ID'].apply(lambda x: x in advertiser_offers))]

        grouped_offers = affiliate_advertiser_drops.groupby('Hitpath Offer ID')[['Clicks','Revenue','Delivered']].sum().reset_index()
        grouped_offers['eCPM'] = grouped_offers['Revenue']*1000 / grouped_offers['Delivered']
        grouped_offers['Weighted CTR'] = grouped_offers['Clicks']*300 / grouped_offers['Delivered']
        grouped_offers['eCPM Rank'] = grouped_offers['eCPM'].rank()
        grouped_offers['CTR Rank'] = grouped_offers['Weighted CTR'].rank()

        grouped_offers['Prob'] = grouped_offers['eCPM']*2 + grouped_offers['Weighted CTR']

        sent_offers = grouped_offers.sort_values(by='Prob',ascending=False)['Hitpath Offer ID'].tolist()
        unsent_offers = lexi[lexi['Hitpath Offer ID'].apply(lambda x: x in advertiser_offers)].groupby('Hitpath Offer ID')[['Clicks','Revenue','Delivered']].sum().reset_index().sort_values(by='Revenue',ascending=False)['Hitpath Offer ID'].tolist()

        return sent_offers+unsent_offers

    def return_unpaused_offers(self, sc_dppub_affiliate, offers_list, cpam_bool=False):
#         import inspect
        offers = self.offers
        emit = self.emit
        affiliate = sc_dppub_affiliate.strip().split('_')[2]

        
        valid_offers = []
        for hitpath in offers_list:

            live_pubids = offers.loc[hitpath]['Live Pub IDs']
            paused_pubids = offers.loc[hitpath]['Paused Pubs']
            
            if not ((type(live_pubids)==float)|pd.isnull(live_pubids)| (live_pubids == '')):

                live_pubids = str(live_pubids).replace('\n',',')
                live_pubids = str(live_pubids).split(',') 
                live_pubids = [x.strip() for x in live_pubids]

                if str(sc_dppub_affiliate) in live_pubids:
                    valid_offers.append(hitpath)

            elif not ((type(paused_pubids)==float)|pd.isnull(paused_pubids)| (paused_pubids == '')):
                paused_pubids = str(paused_pubids).replace('\n',',')
                paused_pubids = str(paused_pubids).split(',') 
                paused_pubids = [x.strip() for x in paused_pubids]
                if str(affiliate) not in paused_pubids:
                    valid_offers.append(hitpath)
            else:
                op_status = offers.loc[hitpath]['Status']
                if 'paused - budget' in op_status.lower():
                    valid_offers.append(hitpath)
                elif 'paused' not in op_status.lower():
                    valid_offers.append(hitpath)
        if not cpam_bool:
            valid_offers = [offer for offer in valid_offers if offers.loc[offer]['Status'] in (['Live','Budgeted Offer - Budget Allocated'])]
        
        #rxp_accounts = emit[~emit['RXP Date'].isna()]['SC_DP&Pub'].unique().tolist()
        da_offers = offers[offers['Scheduling Name'].str.contains('Direct Agents',na=False)].index.tolist()
        #if sc_dppub_affiliate in rxp_accounts:
            #valid_offers = [offer for offer in valid_offers if offer not in da_offers]
        
        return valid_offers


    def find_offer_gap(self,sc_dppub_affiliate,hitpath):
        start = time.time()
        lexi = self.lexi
        fvs = self.fvs
        emit = self.emit
        offers = self.offers
        
        recent_days = lexi['Date'].max() - timedelta(days=30)
        affiliate_recent =  lexi[(lexi['SC_DP&Pub']==sc_dppub_affiliate) & (lexi['Date']>=recent_days) ]
        affiliate_recent_production = affiliate_recent[(affiliate_recent['Send Strategy']=='P') |(affiliate_recent['Segments'].str.contains('DC',na=False))].groupby(['Date','Hitpath Offer ID']).agg({'Revenue':'sum','Delivered':'sum'}).reset_index()

        true_mirrors = self.true_mirrors
        
        if true_mirrors.get(hitpath) is None:
            true_mirrors[hitpath] = [] 

        hit_dates = affiliate_recent_production[(affiliate_recent_production['Hitpath Offer ID']==hitpath) |
                                                (affiliate_recent_production['Hitpath Offer ID'].isin(true_mirrors.get(hitpath)))
                                               ]['Date']
        hit_dates = hit_dates[~hit_dates.duplicated()]

        gap = 7 #default to one week gap
        if len(hit_dates)>=2:
            gap = hit_dates.diff().mean().days

            if gap > 7:
                gap = gap-2


        # Adjust for poor recent performance
        affiliate_recent_production['eCPM'] = affiliate_recent_production['Revenue']*1000 / affiliate_recent_production['Delivered']
        
        account_ecpm = affiliate_recent_production['Revenue'].sum()*1000 / affiliate_recent_production['Delivered'].sum()

        recent_poor_drop_count = len(affiliate_recent_production[(affiliate_recent_production['Hitpath Offer ID']==hitpath)&(affiliate_recent_production['eCPM']<(account_ecpm*0.5))])   
        gap = gap + recent_poor_drop_count*2 #two day gap penalty for every drop below the account's 30 day P average
                
        
        #Adjust Gapping for New Payouts
        hitpath_drops = lexi[(lexi['SC_DP&Pub']==sc_dppub_affiliate)&(lexi['Hitpath Offer ID']==hitpath)]
        if len(hitpath_drops)>0:
            most_recent_drop = hitpath_drops.iloc[-1]
            if most_recent_drop['Payout eCPM'] > most_recent_drop['eCPM']:
                gap = gap - 2
            elif most_recent_drop['Payout eCPM'] < most_recent_drop['eCPM']:
                gap = gap + 2

        pubid = sc_dppub_affiliate.strip()[-6:]
        data_vertical = emit[emit['Affiliate ID']==pubid]['Sub Vertical'].values[0]
        offer_vertical = offers.loc[hitpath]['Vertical']
        if offer_vertical == 'Loan':
            offer_vertical = 'Personal Loan'
        if data_vertical==offer_vertical: # Allow for shorter gaps if similar offer and data verticals
            gap = max(gap,1)
        else:
            gap = max(gap,2) #min gap 2 days
            
        gap = min(gap,10) #max gap 10 days

        end = time.time()  
#         print("\nTotal Time: {} ".format(end-start))
        
        return gap
    

    def find_gap_sample_test(self,sc_dppub_affiliate,hitpath):
        # Test Function for gapping
        lexi = self.lexi
        cobra = self.lexi

        return int(int(sc_dppub_affiliate)/hitpath) % 10 


    def find_mining_sample_test(self,sc_dppub_affiliate):
        # Test Function for mining drops
        lexi = self.lexi

        return [5786,8025,4032,6458]*100

    def find_ot_offers(self,sc_dppub_affiliate):
        #old
        lexi = self.lexi
        emit = self.emit
        fvs = self.fvs

#         find_viper_settings.set_lexi(lexi)
#         find_viper_settings.set_offers(offers)
        test_report = viper_main.get_test_report()
        """ 
        # email version 
        dp_pub = emit[emit['Revenue Pub ID']==int(sc_dppub_affiliate)]['DP&Pub'].mode().values[0]

        if dp_pub in test_report.sheet_names:
            tests = pd.read_excel(test_report,dp_pub)
        else:
            tests = pd.DataFrame({'Hitpath Offer ID' : []})

        test_hitpaths = sorted(tests['Hitpath Offer ID'].tolist(), key = lambda hit : len(fvs.find_days(hit)))
        requests_redrops_list = fvs.find_AM_requests(sc_dppub_affiliate,'test')

        if len(requests_redrops_list)>0:
            print('Adding',str(requests_redrops_list)[1:-1],'to the potential OT list from AM Requests:')

        ot_offers = requests_redrops_list + test_hitpaths
        """
        # sms version 
        self.sc_dppub_affiliate = sc_dppub_affiliate
        sc = self.sc_dppub_affiliate.split('_')[0]
        dppub = self.sc_dppub_affiliate.split('_')[1]
        sc_dppub = self.sc + "_" +self.dppub
        tests = test_report[test_report['shortcode_DP.SV']== sc_dppub]
        test_hitpaths = sorted(tests['Hitpath Offer ID'].tolist(), key = lambda hit : len(fvs.find_days(hit)))
        ot_offers = test_hitpaths
        return ot_offers


    def find_ot_sample_test(self,sc_dppub_affiliate):
        # Test Function for OT drops

        return [5786,4032,8613,5445]


    def find_e_sample_test(self,sc_dppub_affiliate):
        # Test Function for E drops

        return [5786,4032,8613,5445]


    def find_e_offers(self,sc_dppub_affiliate):    
        lexi = self.lexi

        recent_days = lexi['Date'].max() - timedelta(days=30)
        affiliate_recent =  lexi[(lexi['SC_DP&Pub']==sc_dppub_affiliate) & (lexi['Date']>=recent_days) ]
        affiliate_recent_production = affiliate_recent[(affiliate_recent['Send Strategy']=='P')
                                                      |(affiliate_recent['Segment'].str.contains('A120',na=False))]



        high_ctr_offers = (affiliate_recent[(affiliate_recent['Payout Type']!='CPM') &
                              ((affiliate_recent['Send Strategy']=='P') | (affiliate_recent['Send Strategy']=='E') | (affiliate_recent['Send Strategy']=='OT'))  &
                              ((affiliate_recent['Status']=='Live') )
                             ]).sort_values(by='CTR',ascending=False)['Hitpath Offer ID'].unique().tolist()

        return high_ctr_offers

    def find_tier_one_sample_offers(self,sc_dppub_affiliate):
        # Test Function for tier one drops

        return [5786,4032,8613,5445]

    def find_tier_two_sample_offers(self,sc_dppub_affiliate):
        # Test Function for tier two drops

        return [5786,4032,8613,5445]

    def find_tier_three_sample_offers(self,sc_dppub_affiliate):
        # Test Function for three drops

        return [5786,4032,8613,5445]

    def find_tier_two_offers(self, sc_dppub_affiliate):
        cobra = self.cobra
        today = self.today
        
        redrops_dict = self.redrop_ot_report[0]       
        tier_twos = redrops_dict[sc_dppub_affiliate]
        
        upcoming_hitpaths = cobra[(cobra['Dataset']==sc_dppub_affiliate)&
                                  (cobra['Date'].dt.date>=today)]['Hitpath Offer ID'].unique().tolist()
        
        tier_twos = [x for x in tier_twos if x not in upcoming_hitpaths]
        return tier_twos
    
    def find_offer_tests(self, sc_dppub_affiliate):
        cobra = self.cobra
        today = self.today
        
        redrops_dict = self.redrop_ot_report[1]       
        offer_tests = redrops_dict[sc_dppub_affiliate]
        
        upcoming_hitpaths = cobra[(cobra['Dataset']==sc_dppub_affiliate)&
                                  (cobra['Date'].dt.date>=today)]['Hitpath Offer ID'].unique().tolist()
        
        offer_tests = [x for x in offer_tests if x not in upcoming_hitpaths]
        
        return offer_tests

    
    def generate_redrops_and_ots(self,sc_dppub_affiliate='all'):
        lexi_90 = self.lexi_90
        lexi = self.lexi
        today = self.today

        if sc_dppub_affiliate == 'all':
            sc_dppub_affiliates = lexi_90['SC_DP&Pub'].unique()
            generate_report = True
        else:
            sc_dppub_affiliates = [sc_dppub_affiliate]
            generate_report = False

        similar_lists = self.get_similars()
        
        similar_lists_exp = similar_lists.explode('Similar Affiliate IDs').reset_index()
        lexi_90_live = lexi_90[(lexi_90['Send Strategy']=='P') & (lexi_90['Status']=='Live')
                               & (lexi_90['Payout Type']!='CPM')]
        similars = pd.merge(similar_lists_exp, lexi_90_live[['SC_DP&Pub','Date','Hitpath Offer ID','Revenue','Delivered']], how='left', left_on='Similar Affiliate IDs', right_on='SC_DP&Pub')

        similar_offers_agg = similars.groupby(['Affiliate ID Prime','SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'sum','Delivered':'sum','Date':'count'}).reset_index()

        similar_p_ecpms = lexi_90_live.groupby(['SC_DP&Pub']).agg({'Revenue':'sum','Delivered':'sum'})
        similar_p_ecpms['Similar Affiliate P eCPM'] = similar_p_ecpms['Revenue']*1000 / similar_p_ecpms['Delivered']

        similar_offers_agg = pd.merge(similar_offers_agg, similar_p_ecpms[['Similar Affiliate P eCPM']], how='left', on='SC_DP&Pub')

        similar_offers_agg['Similar Affiliate Hitpath eCPM'] = similar_offers_agg['Revenue']*1000 / similar_offers_agg['Delivered']
        similar_offers_agg.rename(columns={'SC_DP&Pub':'Similar Affiliate ID','Date':'Similar Affiliate Hitpath Drop Count','Revenue':'Similar Affiliate Hitpath Revenue','Delivered':'Similar Affiliate Hitpath Delivered'}, inplace=True)

        lexi_30 = lexi_90[lexi_90['Date'].dt.date>=today+timedelta(days=-30)]
        new_df = lexi_30.groupby(['SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'count'}).rename(columns={'Revenue':'Prime Recent Hitpath Count'})
        similar_offers_agg = pd.merge(similar_offers_agg, new_df, how='left', left_on=['Affiliate ID Prime','Hitpath Offer ID'], right_on=['SC_DP&Pub','Hitpath Offer ID'])

        aff_ot_drop_count_df = lexi[lexi['Send Strategy']=='OT'].groupby(['SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'count'}).rename(columns={'Revenue':'Prime All OT Hitpath Count'})
        similar_offers_agg = pd.merge(similar_offers_agg, aff_ot_drop_count_df, how='left', left_on=['Affiliate ID Prime','Hitpath Offer ID'], right_on=['SC_DP&Pub','Hitpath Offer ID'])

        aff_ecpm_group = lexi[(lexi['Send Strategy']=='P')].groupby(['SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'sum','Delivered':'sum'})
        aff_ecpm_group['Prime Affiliate Hitpath P eCPM'] = aff_ecpm_group['Revenue']*1000 / aff_ecpm_group['Delivered']
        similar_offers_agg = pd.merge(similar_offers_agg, aff_ecpm_group[['Prime Affiliate Hitpath P eCPM']], how='left', left_on=['Affiliate ID Prime','Hitpath Offer ID'], right_on=['SC_DP&Pub','Hitpath Offer ID'])

        # similar_lists_exp = similar_lists.explode('Similar Affiliate IDs').reset_index()
        lexi_seasonal_live = lexi[(lexi['Date'].dt.date <= (today + timedelta(days=-330))) & (lexi['Date'].dt.date >= (today + timedelta(days=-390))) & (lexi['Send Strategy']=='P') & (lexi['Status']=='Live')]

        seasonals = pd.merge(similar_lists_exp, lexi_seasonal_live[['SC_DP&Pub','Date','Hitpath Offer ID','Revenue','Delivered']], how='left', left_on='Similar Affiliate IDs', right_on='SC_DP&Pub')

        seasonal_offers_agg = seasonals.groupby(['Affiliate ID Prime','SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'sum','Delivered':'sum','Date':'count'}).reset_index()

        seasonal_p_ecpms = lexi_seasonal_live.groupby(['SC_DP&Pub']).agg({'Revenue':'sum','Delivered':'sum'})
        seasonal_p_ecpms['Seasonal Affiliate P eCPM'] = seasonal_p_ecpms['Revenue']*1000 / seasonal_p_ecpms['Delivered']

        seasonal_offers_agg = pd.merge(seasonal_offers_agg, seasonal_p_ecpms[['Seasonal Affiliate P eCPM']], how='left', on='SC_DP&Pub')
        # seasonal_offers_agg[seasonal_offers_agg['SC_DP&Pub']=='161128']

        seasonal_offers_agg['Seasonal Affiliate Hitpath eCPM'] = seasonal_offers_agg['Revenue']*1000 / seasonal_offers_agg['Delivered']
        seasonal_offers_agg.rename(columns={'SC_DP&Pub':'Seasonal Affiliate ID','Date':'Seasonal Affiliate Hitpath Drop Count','Revenue':'Seasonal Affiliate Hitpath Revenue','Delivered':'Seasonal Affiliate Hitpath Delivered'},inplace=True)
        seasonal_offers_agg[seasonal_offers_agg['Seasonal Affiliate ID']=='161128']

        all_offers_df = pd.merge(similar_offers_agg, seasonal_offers_agg, how='outer', on=['Affiliate ID Prime','Hitpath Offer ID'])
        all_offers_df['Prime Recent Hitpath Count'].fillna(0,inplace=True)
        all_offers_df['Prime All OT Hitpath Count'].fillna(0,inplace=True)
        all_offers_df = all_offers_df.set_index('Affiliate ID Prime')
        
        #Look at company-wide ranking/trends here
        hitpath_ecpm_ranks = lexi_90_live.groupby('Hitpath Offer ID').agg({'eCPM':'mean'}).rank(ascending=False)
        hitpath_ecpm_ranks = hitpath_ecpm_ranks.sort_values('eCPM',ascending=False)
        hitpath_ecpm_ranks = hitpath_ecpm_ranks[:int(len(hitpath_ecpm_ranks)/3):-1] #find top 2/3 of offers by eCPM
        top_p_hitpaths = list(hitpath_ecpm_ranks.index.values)
        all_offers_df['Hitpath in Recent Top P Bool'] = np.where(all_offers_df['Hitpath Offer ID'].isin(top_p_hitpaths),True,False)

        trend_df = lexi_90.groupby(['Hitpath Offer ID',pd.Grouper(key='Date',freq='W')]).agg({'eCPM':'mean'})
        trend_df['eCPM Percent Change'] = trend_df.groupby('Hitpath Offer ID')['eCPM'].pct_change()
        trend_conditions = [
            (trend_df['eCPM Percent Change'] >= 0) ,
            (trend_df['eCPM Percent Change'] < 0 ) ]

        trend_df['eCPM Delta Binary'] = np.select(trend_conditions, [1,-1], default=0)
        trend_overview = trend_df.reset_index().groupby('Hitpath Offer ID').agg({'eCPM Delta Binary':'mean'})
        positive_trending_offers = list(trend_overview['eCPM Delta Binary'][trend_overview['eCPM Delta Binary'] > 0.4].index) #can adjust 0.4 higher for more restrictive
        all_offers_df['Hitpath Trending High'] = np.where(all_offers_df['Hitpath Offer ID'].isin(positive_trending_offers),True,False)

        
        #Get Offer Requests/Suggestions from the last 30 days
        requests_frame = self.viper_settings_sheet.worksheet_by_title('Offer Scale Requests').get_as_df()
        requests_frame['Date Added'] = requests_frame['Date Added'].astype('datetime64[ns]')
        requests = requests_frame[requests_frame['Date Added'].dt.date>=today+timedelta(days=-30)]['Hitpath Offer ID'].tolist()

        affiliate_offer_tests = {}
        affiliate_redrops = {}
        for sc_dppub_affiliate in sc_dppub_affiliates: #come back to update even faster here

            offers_df = all_offers_df.loc[sc_dppub_affiliate]
            
            #Adjust Values for Requested Offers
            offers_df['Requested Bool'] = offers_df['Hitpath Offer ID'].apply(lambda x: x in requests)

            offers_df['Similar Affiliate Hitpath eCPM'] = np.where(offers_df['Requested Bool']==True,
                                                        offers_df['Similar Affiliate Hitpath eCPM']*1.2,
                                                        offers_df['Similar Affiliate Hitpath eCPM'])

            offers_df['Seasonal Affiliate Hitpath eCPM'] = np.where(offers_df['Requested Bool']==True,
                                                        offers_df['Seasonal Affiliate Hitpath eCPM']*1.2,
                                                        offers_df['Seasonal Affiliate Hitpath eCPM'])

            offers_df['Similar Affiliate Hitpath Drop Count'] = np.where(offers_df['Requested Bool']==True,
                                                        offers_df['Similar Affiliate Hitpath Drop Count']+1,
                                                        offers_df['Similar Affiliate Hitpath Drop Count'])

            offers_df['Seasonal Affiliate Hitpath Drop Count'] = np.where(offers_df['Requested Bool']==True,
                                                        offers_df['Seasonal Affiliate Hitpath Drop Count']+1,
                                                        offers_df['Seasonal Affiliate Hitpath Drop Count'])



            #Offers not dropped recently (last 30 days) in the account to schedule
            #Offers that exceeded eCPM P average from a similar account in last 90 days 
            #OR 
            #Offers that exceeded eCPM P average from a similar account or the account to schedule at this time one year ago
            #At least 5 P drops in the last 90 days or at least 3 P drops this time one year ago from similar accounts or the account itself dropped it
            tier_twos = offers_df[(offers_df['Prime Recent Hitpath Count']<1) 
            & ((offers_df['Hitpath in Recent Top P Bool']==True) | (offers_df['Hitpath Trending High']==True))
            & (((offers_df['Similar Affiliate Hitpath eCPM']>offers_df['Similar Affiliate P eCPM']))|(offers_df['Seasonal Affiliate Hitpath eCPM']>offers_df['Seasonal Affiliate P eCPM']))
            & (((offers_df['Similar Affiliate Hitpath Drop Count']>=5))|(offers_df['Seasonal Affiliate Hitpath Drop Count']>=3)|(offers_df['Similar Affiliate ID']==sc_dppub_affiliate))]

            offer_tests = tier_twos[(tier_twos['Prime Affiliate Hitpath P eCPM'].isna())&(tier_twos['Prime All OT Hitpath Count']==0)].sort_values(by=['Requested Bool','Prime Affiliate Hitpath P eCPM'],ascending=[False,False])['Hitpath Offer ID'].unique().tolist()
            redrops = tier_twos[~tier_twos['Prime Affiliate Hitpath P eCPM'].isna()].sort_values(by=['Requested Bool','Prime Affiliate Hitpath P eCPM'],ascending=[False,False])['Hitpath Offer ID'].unique().tolist()

            recent_tests = lexi_30[((lexi_30['Send Strategy']=='PT')|(lexi_30['Send Strategy']=='RT'))&(lexi_30['Status']=='Live')]
            recent_tests = recent_tests.rename(columns={'SC_DP&Pub': 'Similar Affiliate ID'})
            aff_lists = pd.DataFrame(offers_df['Similar Affiliate ID'].dropna().unique()).rename(columns={0:'Similar Affiliate ID'})

            recent_tests = pd.merge(aff_lists, recent_tests[['Similar Affiliate ID','Hitpath Offer ID','eCPM']], how='inner', on='Similar Affiliate ID')
            similar_p_ecpms = similar_p_ecpms.rename_axis('Similar Affiliate ID')

            recent_tests = pd.merge(recent_tests[['Similar Affiliate ID','Hitpath Offer ID','eCPM']], similar_p_ecpms[['Similar Affiliate P eCPM']], how='left', on='Similar Affiliate ID')
            
            #Adjust Values for Requested Offers
            recent_tests['Requested Bool'] = recent_tests['Hitpath Offer ID'].apply(lambda x: x in requests)
            recent_tests['eCPM'] = np.where(recent_tests['Requested Bool']==True,
                                                        recent_tests['eCPM']*1.2,
                                                        recent_tests['eCPM'])
            
            new_ot_tests = recent_tests[recent_tests['eCPM']>recent_tests['Similar Affiliate P eCPM']].groupby('Hitpath Offer ID').agg({'eCPM':'mean','Requested Bool':'sum'}).sort_values(by=['Requested Bool','eCPM'], ascending=[False,False]).index.tolist()


            if generate_report:
                affiliate_offer_tests[sc_dppub_affiliate] = self.return_unpaused_offers(sc_dppub_affiliate, (new_ot_tests + offer_tests))
                affiliate_redrops[sc_dppub_affiliate] = self.return_unpaused_offers(sc_dppub_affiliate, redrops)
            else:
                affiliate_offer_tests[sc_dppub_affiliate] = new_ot_tests + offer_tests
                affiliate_redrops[sc_dppub_affiliate] = redrops
                
        return affiliate_redrops, affiliate_offer_tests
    
    def get_similars(self, list_length=20):
        lexi_90 = self.lexi_90
        affiliates = lexi_90[lexi_90['Date']>=lexi_90['Date'].max()+timedelta(days=-7)]['SC_DP&Pub'].unique()
        
#         affiliates = lexi_90['SC_DP&Pub'].unique()
        

        hits = lexi_90.groupby(['SC_DP&Pub','Hitpath Offer ID']).agg({'Revenue':'sum','Delivered':'sum','Clicks':'sum','Hitpath Offer ID':'count'}).rename(columns={'Hitpath Offer ID':'Hitpath Drop Count'}).reset_index()
        hits['eCPM'] = hits['Revenue']*1000 / hits['Delivered']
        hits['CTR'] = hits['Clicks'] / hits['Delivered']

        hits_grouped = hits.groupby(['SC_DP&Pub'])
        hit_counts = hits_grouped.agg({'Hitpath Offer ID':'count'}).reset_index().rename(columns={'Hitpath Offer ID':'Total Affiliate Hitpath Count'})

        hits = pd.merge(hits, hit_counts, how='left', on='SC_DP&Pub')

        hits['eCPM Rank'] = hits_grouped['eCPM'].rank(ascending=False)
        hits['CTR Rank'] = hits_grouped['CTR'].rank(ascending=False)
        hits['Drop Rank'] = hits_grouped['Hitpath Drop Count'].rank(ascending=False)
        hits['Total Rank'] = hits['eCPM Rank'] + hits['CTR Rank'] + hits['Drop Rank']

        condlist = [hits['Total Rank']<=hits['Total Affiliate Hitpath Count'], ((hits['Total Rank']>hits['Total Affiliate Hitpath Count']) & (hits['Total Rank']<=2*hits['Total Affiliate Hitpath Count']) )]
        hits['Tier'] = np.select(condlist, ['Tier One','Tier Two'], 'Tier Three')

        #Create df with similar sc_dppub_affiliate ids
        affiliates_df = hits[['SC_DP&Pub', 'Tier', 'Hitpath Offer ID']]
        grouped_df = affiliates_df.groupby(['SC_DP&Pub', 'Tier'])['Hitpath Offer ID'].agg(set)
        merged_df = grouped_df.reset_index().merge(grouped_df.reset_index(), on='Tier')
        # merged_df = merged_df[merged_df['Tier']!='Tier Three']

        merged_df['Common Hitpaths'] = merged_df.apply(lambda row: len(row['Hitpath Offer ID_x'].intersection(row['Hitpath Offer ID_y'])), axis=1)
        multiplier_dict = {
            'Tier One': 3,
            'Tier Two': 2,
            'Tier Three': 1
        }

        merged_df['Common Hitpaths Modified'] = merged_df['Tier'].map(multiplier_dict) * merged_df['Common Hitpaths']

        merged_df_summed = merged_df.groupby(['SC_DP&Pub_x','SC_DP&Pub_y']).agg({'Common Hitpaths Modified':'sum'})

        #This will potentially cut off some sc_dppub_affiliate ids if they have a tie score...tiebreaker goes to older Affiliate
        similar_lists = merged_df_summed.reset_index().groupby('SC_DP&Pub_x').apply(lambda x: x.sort_values(['Common Hitpaths Modified','SC_DP&Pub_y'], ascending=[False,True])['SC_DP&Pub_y'].tolist()[:list_length+1]).reset_index(name='Similar Affiliate IDs')
        similar_lists = similar_lists.rename(columns={'SC_DP&Pub_x':'Affiliate ID Prime'}).set_index('Affiliate ID Prime')
        return similar_lists
    
    
    def get_similars_verticals(self, list_length=20):
        #Return a df with the top similar accounts for each account (note that the account itself will be the first entry in each list and so the actual list length will be one longer than "list_length"
        lexi_90 = self.lexi_90
        affiliates = lexi_90['SC_DP&Pub'].unique()

        verts = lexi_90.groupby(['SC_DP&Pub','Vertical']).agg({'Revenue':'sum','Delivered':'sum','Clicks':'sum','Vertical':'count'}).rename(columns={'Vertical':'Vertical Drop Count'}).reset_index()
        verts['eCPM'] = verts['Revenue']*1000 / verts['Delivered']
        verts['CTR'] = verts['Clicks'] / verts['Delivered']

        verts_grouped = verts.groupby(['SC_DP&Pub'])
        vert_counts = verts_grouped.agg({'Vertical':'count'}).reset_index().rename(columns={'Vertical':'Total Affiliate Vertical Count'})
        verts = pd.merge(verts, vert_counts, how='left', on='SC_DP&Pub')

        verts['eCPM Rank'] = verts_grouped['eCPM'].rank(ascending=False)
        verts['CTR Rank'] = verts_grouped['CTR'].rank(ascending=False)
        verts['Drop Rank'] = verts_grouped['Vertical Drop Count'].rank(ascending=False)
        verts['Total Rank'] = verts['eCPM Rank'] + verts['CTR Rank'] + verts['Drop Rank']

        condlist = [verts['Total Rank']<=verts['Total Affiliate Vertical Count'], ((verts['Total Rank']>verts['Total Affiliate Vertical Count']) & (verts['Total Rank']<=2*verts['Total Affiliate Vertical Count']) )]
        verts['Tier'] = np.select(condlist, ['Tier One','Tier Two'], 'Tier Three')

        #Create df with similar sc_dppub_affiliate ids
        affiliates_df = verts[['SC_DP&Pub', 'Tier', 'Vertical']]
        grouped_df = affiliates_df.groupby(['SC_DP&Pub', 'Tier'])['Vertical'].agg(set)
        merged_df = grouped_df.reset_index().merge(grouped_df.reset_index(), on='Tier')
        merged_df['Common Verticals'] = merged_df.apply(lambda row: len(row['Vertical_x'].intersection(row['Vertical_y'])), axis=1)
        merged_df.sort_values(by=['Affiliate ID_x','Common Verticals'],ascending=[True,False])
        merged_df_summed = merged_df.groupby(['Affiliate ID_x','Affiliate ID_y']).agg({'Common Verticals':'sum'})

        #This will potentially cut off some sc_dppub_affiliate ids if they have a tie score...tiebreaker goes to older Affiliate
        similar_lists = merged_df_summed.reset_index().groupby('Affiliate ID_x').apply(lambda x: x.sort_values(['Common Verticals','Affiliate ID_y'], ascending=[False,True])['Affiliate ID_y'].tolist()[:list_length+1]).reset_index(name='Similar Affiliate IDs')
        similar_lists = similar_lists.rename(columns={'Affiliate ID_x':'Affiliate ID Prime'}).set_index('Affiliate ID Prime')

        return similar_lists
    
    def generate_full_report(self):
        import time
        start = time.time()
        affiliate_redrops, affiliate_offer_tests = self.generate_redrops_and_ots('all')
        end = time.time()
        print("Generation Time:",round((end-start),2),"seconds.")
     
        
#         empty_ot_keys = [k for k, v in affiliate_offer_tests.items() if len(v) == 0]
#         for empty_ot_key in empty_ot_keys:
#             affiliate_offer_tests.pop(empty_ot_key)
#             print('No Offer Tests Found for {}!'.format(empty_ot_key))
            
#         empty_redrop_keys = [k for k, v in affiliate_redrops.items() if len(v) == 0]
#         for empty_redrop_key in empty_redrop_keys:
#             affiliate_redrops.pop(empty_redrop_key)
#             print('No Redrops Found for {}!'.format(empty_redrop_key))
            
    
        offers = self.offers
        cs = self.cs
        fvs = self.fvs
        
        # Make DataFrames with More Information
        ot_df = pd.DataFrame(affiliate_offer_tests.items(), columns=['SC_DP&Pub','Hitpath Offer ID']).explode('Hitpath Offer ID').reset_index().drop('index',axis=1)
#         offer_names = offers.groupby('Hitpath Offer ID')[['Scheduling Name','Day Restrictions']].agg(pd.Series.mode)
        ot_df = pd.merge(ot_df, offers[['Scheduling Name','Day Restrictions']], how='left', left_on='Hitpath Offer ID', right_on=offers.index)
        ot_df['Day Restrictions'].replace(np.nan, 'None', inplace=True)
        ot_df['Day Restrictions'] = ot_df['Day Restrictions'].apply(lambda y: 'None' if len(y)==0 else y)
#         ot_df['Suggested Creative'] = ot_df.apply(lambda x: cs.choose_creative(x['SC_DP&Pub'], x['Hitpath Offer ID'], simple=True, find_id=False)[0] if not pd.isna(x['Hitpath Offer ID']) else np.nan,axis=1)
        ot_df['Suggested Creative'] = ot_df.apply(lambda x: cs.lookup_creative_choice(x['SC_DP&Pub'], x['Hitpath Offer ID']) if not pd.isna(x['Hitpath Offer ID']) else np.nan,axis=1)
#         ot_df['Suggested Creative'] = 'Test'
        ot_df.set_index(['SC_DP&Pub','Hitpath Offer ID'], inplace=True)
        ot_df = ot_df[~ot_df.index.duplicated(keep='first')]
        

#         print('Starting redrops')
        redrop_df = pd.DataFrame(affiliate_redrops.items(), columns=['SC_DP&Pub','Hitpath Offer ID']).explode('Hitpath Offer ID').reset_index().drop('index',axis=1)
        # offer_names = offers.groupby('Hitpath Offer ID')[['Scheduling Name','Day Restrictions']].agg(pd.Series.mode)
        redrop_df = pd.merge(redrop_df, offers[['Scheduling Name','Day Restrictions']], how='left', left_on='Hitpath Offer ID', right_on=offers.index)
        redrop_df['Day Restrictions'].replace(np.nan, 'None', inplace=True)
        redrop_df['Day Restrictions'] = redrop_df['Day Restrictions'].apply(lambda y: 'None' if len(y)==0 else y)
#         redrop_df['Suggested Creative'] = redrop_df.apply(lambda x: cs.choose_creative(x['SC_DP&Pub'], x['Hitpath Offer ID'], simple=True, find_id=False)[0] if not pd.isna(x['Hitpath Offer ID']) else np.nan,axis=1)
        redrop_df['Suggested Creative'] = redrop_df.apply(lambda x: cs.lookup_creative_choice(x['SC_DP&Pub'], x['Hitpath Offer ID']) if not pd.isna(x['Hitpath Offer ID']) else np.nan,axis=1)
#         redrop_df['Suggested Creative'] = 'Testing'
        redrop_df.set_index(['SC_DP&Pub','Hitpath Offer ID'], inplace=True)
        redrop_df = redrop_df[~redrop_df.index.duplicated(keep='first')]
   
        # Write to xlsx file
        from openpyxl import Workbook
        from openpyxl.styles import Font, Color, colors, fills
        from openpyxl.styles.borders import Border, Side, BORDER_THIN
        thin_border = Border(
            left=Side(border_style=BORDER_THIN, color='00616161'),
            right=Side(border_style=BORDER_THIN, color='00616161'),
            top=Side(border_style=BORDER_THIN, color='00616161'),
            bottom=Side(border_style=BORDER_THIN, color='00616161')
        )

        workbook = Workbook()
        ot_sheet = workbook.active
        ot_sheet.title = 'Offer Test Suggestions'

        ot_sheet["A1"] = "Affiliate"
        ot_sheet["B1"] = "Hitpath Offer ID"
        ot_sheet["C1"] = "Scheduling Name"
        ot_sheet["D1"] = "Day Restrictions"
        ot_sheet["E1"] = "Suggested Creative Type"
        ot_sheet["F1"] = "Mirrors Also Suggested"

        mirrors = self.mirrors
        true_mirrors = self.true_mirrors
        all_mirrors = {**mirrors, **true_mirrors}
        
#         colors = ['ffff99', 'a5c4e0', 'e5ccff','ffcce5', 'ffcc99', 'bff5cc', 'fee1d3', 'ccffff','fa7da1', 'c0c0c0', '00cccc'
#                  'ff9999','ff9933','0080ff','00cc00']
        
        colors = ['e0e0e0', 'ffff99', '99ff99', '99ffff', '9999ff', 'ff99ff', 'ff9999',
                  'ffcc99', 'ccff99', '99ffcc', '99ccff', 'cc99ff', 'ff99cc',
                  'ff4444', 'ffff44', '33ff44', '44ffff', 'ff33ff', 'ff44ff', 'a0a0a0',
                  'ffcc44', 'ccff44', '44ffcc', '44ccff', 'cc44ff', 'ff44cc',
                  'ff1111', 'ffff11', '11ff11', '11ffff', '1111ff', 'ff11ff', '808080',
                  'ffcc11', 'ccff11', '11ffcc', '11ccff', 'cc11ff', 'ff11cc']

        row = 2
        start_merge = 2
        rows_to_color = []
        color_assignments = {}
        color_counter = 0
            
        sorted_ots = {aff: hits for aff,hits in sorted(affiliate_offer_tests.items())}
        for sc_dppub_affiliate, suggestions in sorted_ots.items():
            if len(suggestions)>0:
                for offer in suggestions:
                    ot_sheet[f"A{row}"] = sc_dppub_affiliate
                    ot_sheet[f"B{row}"] = offer
                    ot_sheet[f"C{row}"] = ot_df.loc[sc_dppub_affiliate].loc[offer]['Scheduling Name']
                    ot_sheet[f"D{row}"] = ot_df.loc[sc_dppub_affiliate].loc[offer]['Day Restrictions']
                    ot_sheet[f"E{row}"] = ot_df.loc[sc_dppub_affiliate].loc[offer]['Suggested Creative']
                    if all_mirrors.get(offer) is not None:
                        aff_mirrors = set(all_mirrors.get(offer)).intersection(set(ot_df.loc[sc_dppub_affiliate].index))
                        ot_sheet[f"F{row}"] = str(list(aff_mirrors)).replace('[','').replace(']','')
                        
                        if color_assignments.get(offer) is None:
                            color_assignments[offer] = color_counter
                            for mirror in all_mirrors.get(offer):
                                color_assignments[mirror] = color_counter
                            color_counter+=1
                        color = colors[color_assignments[offer]]

                        if len(aff_mirrors)>0:
                            for cell in ot_sheet[(str(row)):(str(row))][1:]:
                                cell.fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb=color))
                                cell.border = thin_border
                                
                    row+=1
                ot_sheet.merge_cells(start_row=start_merge, start_column = 1, end_row=row-1, end_column=1)
            else:
                ot_sheet[f"A{row}"] = sc_dppub_affiliate
                ot_sheet[f"B{row}"] = offer
                ot_sheet[f"C{row}"] = 'No Suitable Offer Tests Found!'
                ot_sheet[f"D{row}"] = ''
                ot_sheet[f"E{row}"] = ''
                row+=1
                ot_sheet.merge_cells(start_row=start_merge, start_column = 1, end_row=row-1, end_column=1)
                print('No Offer Tests Found for {}!'.format(sc_dppub_affiliate))
            start_merge = row


        redrop_sheet = workbook.create_sheet('Redrop Suggestions')
        redrop_sheet["A1"] = "Affiliate"
        redrop_sheet["B1"] = "Hitpath Offer ID"
        redrop_sheet["C1"] = "Scheduling Name"
        redrop_sheet["D1"] = "Day Restrictions"
        redrop_sheet["E1"] = "Suggested Creative Type"
        redrop_sheet["F1"] = "Mirrors Also Suggested"

        row = 2
        start_merge = 2
        sorted_redrops = {aff: hits for aff,hits in sorted(affiliate_redrops.items())}
        for sc_dppub_affiliate, suggestions in sorted_redrops.items():
            if len(suggestions)>0:
                for offer in suggestions:
                    redrop_sheet[f"A{row}"] = sc_dppub_affiliate
                    redrop_sheet[f"B{row}"] = offer
                    redrop_sheet[f"C{row}"] = redrop_df.loc[sc_dppub_affiliate].loc[offer]['Scheduling Name']
                    redrop_sheet[f"D{row}"] = redrop_df.loc[sc_dppub_affiliate].loc[offer]['Day Restrictions']
                    redrop_sheet[f"E{row}"] = redrop_df.loc[sc_dppub_affiliate].loc[offer]['Suggested Creative']
                    if mirrors.get(offer) is not None:
                        aff_mirrors = set(mirrors.get(offer)).intersection(set(redrop_df.loc[sc_dppub_affiliate].index))
                        redrop_sheet[f"F{row}"] = str(list(aff_mirrors)).replace('[','').replace(']','')
                        
                        if color_assignments.get(offer) is None:
                            color_assignments[offer] = color_counter
                            for mirror in mirrors.get(offer):
                                color_assignments[mirror] = color_counter
                            color_counter+=1
                        color = colors[color_assignments[offer]]
                        
                        if len(aff_mirrors)>0:
                            for cell in redrop_sheet[(str(row)):(str(row))][1:]:
                                cell.fill = fills.PatternFill(patternType='solid', fgColor=Color(rgb=color))
                                cell.border = thin_border
                    row+=1
                redrop_sheet.merge_cells(start_row=start_merge, start_column = 1, end_row=row-1, end_column=1)
            else:
                redrop_sheet[f"A{row}"] = sc_dppub_affiliate
                redrop_sheet[f"B{row}"] = offer
                redrop_sheet[f"C{row}"] = 'No Suitable Redrop Suggestions Found!'
                redrop_sheet[f"D{row}"] = ''
                redrop_sheet[f"E{row}"] = ''
                row+=1
                redrop_sheet.merge_cells(start_row=start_merge, start_column = 1, end_row=row-1, end_column=1)
                print('No Redrops Found for {}!'.format(sc_dppub_affiliate))
            start_merge = row


        all_sheets = [ot_sheet, redrop_sheet]

        for ws in all_sheets:
            dims = {}
            for row in ws.rows:
                for cell in ws["1:1"]:
                    cell.font = Font(bold=True)
                for cell in row:
                    newline_count = 1
                    if cell.value:
                        if type(cell.value)==str:
                            if ('\n' in cell.value):
                                newline_count = cell.value.count('\n')
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))/newline_count 
            for col, value in dims.items():
                ws.column_dimensions[col].width = value

        workbook.save("OT and Redrop Report.xlsx")


